"""
TalaMala v4 - Production Database Seeder
=========================================
Seeds the database with REAL production data only. No test data.

Usage:
    python scripts/seed_production.py          # Seed (idempotent)
    python scripts/seed_production.py --reset  # Drop all + reseed

What this creates:
  - 1 admin user (super_admin)
  - System settings (gold price, tax, etc.)
  - 4 product categories
  - 4 card designs + 4 package types
  - 1 real batch (BATCH-20241203)
  - 44 products (from Excel wages + _private images)
  - 4 dealer tiers + tier wages
  - 31 provinces + cities + districts (Iran geo data)
  - 561 real bars (from _private/bars_data.json)
  - 2 real customers (sold bar owners)

What this does NOT create:
  - No test dealers, customers, coupons, bars, tickets, wallets
"""

import sys
import os
import io
import json
import shutil
import random
import string
from datetime import datetime, timezone
import openpyxl

# Fix Windows console encoding for Persian text
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config.database import SessionLocal, Base, engine
from modules.user.models import User
from modules.admin.models import SystemSetting
from modules.customer.address_models import GeoProvince, GeoCity, GeoDistrict, CustomerAddress
from modules.catalog.models import (
    ProductCategory, ProductCategoryLink, Product, ProductImage, CardDesign, CardDesignImage,
    PackageType, PackageTypeImage, Batch, BatchImage, ProductTierWage,
)
from modules.inventory.models import (
    Bar, BarImage, OwnershipHistory, BarStatus,
    DealerTransfer, BarTransfer, TransferType,
    ReconciliationSession, ReconciliationItem,
    CustodialDeliveryRequest,
)
from modules.cart.models import Cart, CartItem
from modules.order.models import Order, OrderItem, OrderStatusLog
from modules.wallet.models import Account, LedgerEntry, WalletTopup, WithdrawalRequest
from modules.coupon.models import Coupon, CouponMobile, CouponUsage, CouponCategory
from modules.dealer.models import DealerTier, DealerSale, BuybackRequest, SubDealerRelation, B2BOrder, B2BOrderItem
from modules.ticket.models import Ticket, TicketMessage, TicketAttachment, TicketStatus, TicketPriority, TicketCategory, SenderType
from modules.review.models import Review, ReviewImage, ProductComment, CommentImage, CommentLike
from modules.dealer_request.models import DealerRequest, DealerRequestAttachment
from modules.pricing.models import Asset, GOLD_18K, SILVER


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def ensure_tables():
    print("[0] Ensuring all tables exist...")
    Base.metadata.create_all(bind=engine)
    print("  + All tables OK\n")


def ensure_schema_updates():
    """Add missing columns to existing tables (create_all won't do this)."""
    from sqlalchemy import text, inspect
    print("[0.5] Checking schema updates...")

    insp = inspect(engine)
    updates = 0

    # --- products: description ---
    if "products" in insp.get_table_names():
        prod_cols = {c["name"] for c in insp.get_columns("products")}
        with engine.begin() as conn:
            if "description" not in prod_cols:
                conn.execute(text("ALTER TABLE products ADD COLUMN description TEXT"))
                print("  + products.description added")
                updates += 1

    if updates == 0:
        print("  = schema up to date")
    print()


def seed():
    db = SessionLocal()
    try:
        print("=" * 50)
        print("  TalaMala v4 — Production Seeder")
        print("=" * 50)

        ensure_tables()
        ensure_schema_updates()

        # ==========================================
        # 1. Admin User (super_admin only)
        # ==========================================
        print("\n[1] Admin Users")

        admins_data = [
            {"mobile": "09120725564", "full_name": "شکیبا غلامی", "role": "admin"},
            {"mobile": "09121023589", "full_name": "مدیر سیستم", "role": "admin"},
            {"mobile": "09121058447", "full_name": "سید علی حسینی", "role": "admin"},
        ]
        for admin_data in admins_data:
            existing = db.query(User).filter(User.mobile == admin_data["mobile"]).first()
            if not existing:
                parts = admin_data["full_name"].split()
                db.add(User(
                    mobile=admin_data["mobile"],
                    first_name=parts[0] if parts else "",
                    last_name=" ".join(parts[1:]) if len(parts) > 1 else "",
                    is_admin=True,
                    is_customer=True,
                    admin_role=admin_data["role"],
                ))
                print(f"  + admin: {admin_data['mobile']}")
            else:
                print(f"  = exists: {admin_data['mobile']}")

        db.flush()

        # ==========================================
        # 2. System Settings
        # ==========================================
        print("\n[2] System Settings")

        settings_data = {
            "site_name":            ("طلاملا", "نام سایت"),
            "site_logo":            ("", "لوگوی سایت"),
            "support_phone":        ("02112345678", "شماره پشتیبانی"),
            "support_telegram":     ("@talamala", "تلگرام پشتیبانی"),
            "tax_percent":          ("10", "درصد مالیات بر ارزش افزوده"),
            "min_order_amount":     ("10000000", "حداقل مبلغ سفارش (ریال)"),
            "reservation_minutes":  ("15", "مدت زمان رزرو (دقیقه)"),
            "shipping_cost":        ("500000", "هزینه ارسال پستی (ریال)"),
            "insurance_percent":    ("1.5", "درصد بیمه پست"),
            "insurance_cap":        ("500000000", "سقف بیمه پست (ریال)"),
            "gold_spread_percent":  ("2", "اسپرد تبدیل ریال به طلا (درصد) — deprecated"),
            "gold_fee_customer_percent": ("2", "کارمزد خرید/فروش طلا — مشتری عادی (%)"),
            "gold_fee_dealer_percent":   ("0.5", "کارمزد خرید/فروش طلا — نماینده (%)"),
            "silver_fee_customer_percent": ("1.5", "کارمزد خرید/فروش نقره — مشتری عادی (%)"),
            "silver_fee_dealer_percent":   ("0.3", "کارمزد خرید/فروش نقره — نماینده (%)"),
            "active_gateway":       ("zibal", "درگاه پرداخت فعال (zibal/sepehr/top/parsian)"),
            "log_retention_days":   ("45", "مدت نگهداری لاگ درخواست‌ها (روز)"),
        }

        for key, (value, desc) in settings_data.items():
            existing = db.query(SystemSetting).filter(SystemSetting.key == key).first()
            if not existing:
                db.add(SystemSetting(key=key, value=value, description=desc))
                print(f"  + {key} = {value}")
            else:
                print(f"  = exists: {key}")

        db.flush()

        # ==========================================
        # 2b. Assets (Gold / Silver price tracking)
        # ==========================================
        print("\n[2b] Assets")
        if not db.query(Asset).filter(Asset.asset_code == GOLD_18K).first():
            db.add(Asset(
                asset_code=GOLD_18K,
                asset_label="طلای ۱۸ عیار",
                price_per_gram=0,
                stale_after_minutes=15,
                auto_update=True,
                update_interval_minutes=5,
                source_url="https://goldis.ir/price/api/v1/price/assets/gold18k/final-prices",
            ))
            print("  + Asset: gold_18k (auto_update=True)")
        else:
            print("  = exists: gold_18k")

        if not db.query(Asset).filter(Asset.asset_code == SILVER).first():
            db.add(Asset(
                asset_code=SILVER,
                asset_label="نقره",
                price_per_gram=0,
                stale_after_minutes=30,
                auto_update=False,
                update_interval_minutes=10,
            ))
            print("  + Asset: silver (auto_update=False)")
        else:
            print("  = exists: silver")

        db.flush()

        # ==========================================
        # 3. Card Designs + Package Types + Batch
        # ==========================================
        print("\n[3] Catalog Accessories")

        for name in ["طرح کلاسیک", "طرح مدرن", "طرح هدیه", "طرح نوروز"]:
            if not db.query(CardDesign).filter(CardDesign.name == name).first():
                db.add(CardDesign(name=name))
                print(f"  + Design: {name}")

        pkg_data = [
            {"name": "جعبه استاندارد", "price": 0},
            {"name": "جعبه لوکس", "price": 5_000_000},
            {"name": "جعبه هدیه ویژه", "price": 15_000_000},
            {"name": "پاکت ساده", "price": 1_000_000},
        ]
        for pd in pkg_data:
            existing = db.query(PackageType).filter(PackageType.name == pd["name"]).first()
            if not existing:
                db.add(PackageType(name=pd["name"], price=pd["price"], is_active=True))
                print(f"  + Package: {pd['name']} ({pd['price']} rial)")
            else:
                existing.price = pd["price"]

        # Real batch from old system
        batch_data = {"batch_number": "BATCH-20241203", "melt_number": "MELT-001", "operator": "محمد شمسی پور"}
        if not db.query(Batch).filter(Batch.batch_number == batch_data["batch_number"]).first():
            db.add(Batch(**batch_data))
            print(f"  + Batch: {batch_data['batch_number']}")

        db.flush()

        # ==========================================
        # 4. Product Categories
        # ==========================================
        print("\n[4] Product Categories")

        categories_data = [
            ("شمش طلا طلاملا", "gold-talamala", 1),
            ("شمش طلا سرمایه‌ای", "gold-investment", 2),
            ("شمش نقره طلاملا", "silver-talamala", 3),
            ("شمش نقره سرمایه‌ای", "silver-investment", 4),
        ]

        cat_map = {}
        for cat_name, cat_slug, sort_order in categories_data:
            existing = db.query(ProductCategory).filter(ProductCategory.slug == cat_slug).first()
            if not existing:
                cat = ProductCategory(name=cat_name, slug=cat_slug, sort_order=sort_order, is_active=True)
                db.add(cat)
                db.flush()
                cat_map[cat_slug] = cat
                print(f"  + {cat_name}")
            else:
                cat_map[cat_slug] = existing
                print(f"  = exists: {cat_name}")

        db.flush()

        # ==========================================
        # 5. Products (from Excel wages + _private images)
        # ==========================================
        IMG_SRC_BASE = os.path.join(PROJECT_ROOT, "_private", "عکس محصولات")
        IMG_DST_BASE = os.path.join(PROJECT_ROOT, "static", "uploads", "products")
        os.makedirs(IMG_DST_BASE, exist_ok=True)

        # ---- Product descriptions (per weight × metal) ----
        gold_descriptions = {
            0.1: (
                "شمش ۱۰۰ سوتی؛ کوچک‌ترین واحد سرمایه‌گذاری طلا\n\n"
                "شمش ۱۰۰ سوتی (یک‌دهم گرم) نقطه شروعی ایده‌آل برای ورود به بازار طلاست. "
                "این وزن سبک امکان خرید مکرر و پس‌انداز تدریجی را فراهم می‌کند و به‌ویژه "
                "برای هدیه‌های کوچک و نمادین گزینه‌ای بی‌نظیر است.\n\n"
                "هدیه‌ای نمادین: مناسب برای تولد، جشن فارغ‌التحصیلی و مناسبت‌های خاص.\n"
                "پس‌انداز گام‌به‌گام: خرید ماهانه و ساخت سبد طلای شخصی.\n"
                "نقدشوندگی بالا: تبدیل آسان به وجه نقد در هر زمان.\n\n"
                "هر قطعه با خلوص ۹۹۵ (۲۴ عیار) و کد رهگیری اختصاصی عرضه می‌شود "
                "تا اصالت سرمایه شما تضمین گردد."
            ),
            0.2: (
                "شمش ۲۰۰ سوتی؛ قدمی مطمئن در مسیر سرمایه‌گذاری\n\n"
                "شمش ۲۰۰ سوتی (دو‌دهم گرم) تعادلی مناسب بین قیمت پایین و ارزش واقعی طلا "
                "ایجاد می‌کند. این وزن برای کسانی که می‌خواهند به‌صورت منظم طلا پس‌انداز کنند، "
                "انتخابی هوشمندانه است.\n\n"
                "سرمایه‌گذاری تدریجی: امکان خرید مکرر بدون فشار مالی.\n"
                "هدیه‌ای ارزشمند: مناسب برای یادبودها و قدردانی.\n"
                "قابلیت جمع‌آوری: ساخت مجموعه طلای شخصی با خریدهای کوچک.\n\n"
                "با خلوص ۹۹۵ و بسته‌بندی امن، هر شمش دارای شناسه یکتا و گواهی اصالت است."
            ),
            0.5: (
                "شمش نیم گرمی؛ تعادل هوشمندانه ارزش و دسترسی\n\n"
                "شمش ۵۰۰ سوتی (نیم گرم) یکی از محبوب‌ترین وزن‌ها در بازار طلای خرد است. "
                "این وزن، نقطه تلاقی مناسبی بین سرمایه‌گذاری معنادار و قیمت قابل دسترس ایجاد کرده است.\n\n"
                "هدیه‌ای ماندگار: گزینه‌ای برازنده برای عروسی، تولد و مناسبت‌های رسمی.\n"
                "پس‌انداز هدفمند: روشی اصولی برای تبدیل درآمد ماهانه به سرمایه پایدار.\n"
                "نقدشوندگی سریع: فروش آسان در هر زمان با قیمت روز بازار.\n\n"
                "هر قطعه با خلوص ۹۹۵ (۲۴ عیار)، کد رهگیری و بسته‌بندی ایمن عرضه می‌شود."
            ),
            1.0: (
                "شمش یک گرمی؛ جایگاه ویژه در سبد سرمایه‌گذاری\n\n"
                "چرا شمش یک گرمی جایگاه ویژه‌ای در سبد سرمایه‌گذاری دارد؟ وزن یک گرم، "
                "تعادلی هوشمندانه میان نقدشوندگی سریع و قدرت خرید مجدد در دراز مدت ایجاد کرده است.\n\n"
                "هدیه‌ای ماندگار: جایگزینی برازنده برای وجه نقد در مناسبت‌های رسمی و خانوادگی.\n"
                "پس‌انداز مستمر: روشی اصولی برای تبدیل درآمدهای جاری به سرمایه‌ای پایدار.\n"
                "سرمایه‌گذاری امن: مسیری مطمئن برای ورود به بازار طلا، فارغ از ریسک‌های مرسوم.\n\n"
                "خلوص ۹۹۵ (۲۴ عیار): شمش‌های ما با خلوص ۹۹۵ عرضه می‌شوند. "
                "این یعنی شما صاحب خالص‌ترین شکل ممکن طلا هستید.\n"
                "هر قطعه با کد رهگیری اختصاصی و استاندارد تولید، اصالت آن تضمین شده است."
            ),
            2.5: (
                "شمش ۲.۵ گرمی؛ انتخاب حرفه‌ای سرمایه‌گذاران\n\n"
                "شمش ۲.۵ گرمی نقطه عطفی در سرمایه‌گذاری طلا محسوب می‌شود. این وزن، "
                "ضمن حفظ قابلیت نقدشوندگی بالا، ارزش قابل توجهی را در قطعه‌ای فشرده و قابل نگهداری ارائه می‌دهد.\n\n"
                "سرمایه‌گذاری متوسط: مناسب برای کسانی که فراتر از خریدهای خُرد، به دنبال ذخیره ارزش هستند.\n"
                "هدیه‌ای لوکس: انتخابی شایسته برای مناسبت‌های مهم و رویدادهای خاص.\n"
                "پرتفوی متنوع: امکان ترکیب با سایر وزن‌ها برای مدیریت بهتر سبد سرمایه.\n\n"
                "با خلوص ۹۹۵ و گواهی اصالت، هر شمش تحت نظارت دقیق تولید و بسته‌بندی می‌شود."
            ),
            5.0: (
                "شمش ۵ گرمی؛ ذخیره ارزش در ابعاد فشرده\n\n"
                "شمش ۵ گرمی ترکیبی ایده‌آل از ارزش بالا و ابعاد قابل مدیریت است. "
                "این وزن یکی از پرطرفدارترین انتخاب‌ها در میان سرمایه‌گذاران جدی بازار طلاست.\n\n"
                "ذخیره ارزش: نگهداری سرمایه قابل توجه در فضایی بسیار کوچک.\n"
                "نقدشوندگی مناسب: فروش سریع و آسان در بازار با تقاضای بالا.\n"
                "هدیه سرمایه‌ای: انتخابی ممتاز برای هدایای عروسی و رویدادهای مهم زندگی.\n\n"
                "هر شمش با خلوص ۹۹۵ (۲۴ عیار)، کد رهگیری یکتا و بسته‌بندی ایمن ارائه می‌شود."
            ),
            10.0: (
                "شمش ۱۰ گرمی؛ سرمایه‌گذاری قدرتمند و نقدشونده\n\n"
                "شمش ۱۰ گرمی یکی از استانداردترین واحدهای سرمایه‌گذاری در بازار جهانی طلاست. "
                "این وزن، تعادل بهینه‌ای بین حجم سرمایه‌گذاری و سهولت معامله ایجاد می‌کند.\n\n"
                "استاندارد جهانی: وزنی شناخته‌شده و مورد اعتماد در تمام بازارهای طلا.\n"
                "سرمایه‌گذاری کلان: مناسب برای حفظ و رشد سرمایه‌های بزرگ‌تر.\n"
                "نقدشوندگی بالا: قابلیت فروش سریع با کمترین اختلاف قیمت خرید و فروش.\n\n"
                "تولید شده با خلوص ۹۹۵، دارای شناسه یکتا و گواهی اصالت معتبر."
            ),
            20.0: (
                "شمش ۲۰ گرمی؛ گنجینه‌ای برای آینده\n\n"
                "شمش ۲۰ گرمی انتخابی استراتژیک برای سرمایه‌گذاران بلندمدت است. "
                "این وزن ارزش قابل توجهی را در قالبی فشرده و امن ارائه می‌دهد و "
                "گزینه‌ای عالی برای تنوع‌بخشی به سبد دارایی است.\n\n"
                "سرمایه‌گذاری بلندمدت: حفظ ارزش دارایی در برابر تورم و نوسانات اقتصادی.\n"
                "ذخیره امن: نگهداری حجم بالای سرمایه در فضایی بسیار کوچک.\n"
                "نقدشوندگی تضمین‌شده: فروش آسان در بازار با تقاضای همیشگی.\n\n"
                "با خلوص ۹۹۵ (۲۴ عیار) و بسته‌بندی ایمن، اصالت هر قطعه با کد رهگیری تضمین می‌شود."
            ),
            31.1: (
                "شمش یک اونسی؛ استاندارد طلایی بازار جهانی\n\n"
                "شمش یک اونس (۳۱.۱ گرم) معیار اصلی قیمت‌گذاری طلا در بازارهای بین‌المللی است. "
                "خرید شمش اونسی یعنی سرمایه‌گذاری مطابق با استانداردهای جهانی.\n\n"
                "مرجع قیمت جهانی: قیمت طلا در بازارهای بین‌المللی بر اساس اونس تعیین می‌شود.\n"
                "سرمایه‌گذاری حرفه‌ای: انتخاب اول سرمایه‌گذاران حرفه‌ای و نهادهای مالی.\n"
                "نقدشوندگی بی‌رقیب: بالاترین سطح تقاضا و سهولت معامله.\n\n"
                "هر شمش اونسی با خلوص ۹۹۵ (۲۴ عیار)، شناسه یکتا و گواهی اصالت عرضه می‌شود."
            ),
            50.0: (
                "شمش ۵۰ گرمی؛ سرمایه‌گذاری سنگین با حداکثر بازدهی\n\n"
                "شمش ۵۰ گرمی برای سرمایه‌گذارانی طراحی شده که به دنبال ذخیره حجم بالای ارزش هستند. "
                "این وزن، اجرت ساخت کمتری نسبت به وزن‌های سبک‌تر دارد و بازدهی سرمایه‌گذاری را بهبود می‌بخشد.\n\n"
                "بهینه‌سازی هزینه: اجرت ساخت کمتر به نسبت وزن.\n"
                "سرمایه‌گذاری کلان: مناسب برای حفظ ثروت نسل‌ها.\n"
                "نقدشوندگی: قابلیت فروش در بازار حرفه‌ای طلا.\n\n"
                "تولید شده با خلوص ۹۹۵ (۲۴ عیار)، دارای کد رهگیری و بسته‌بندی ایمن ویژه."
            ),
            100.0: (
                "شمش ۱۰۰ گرمی؛ اوج سرمایه‌گذاری در طلای خالص\n\n"
                "شمش ۱۰۰ گرمی بالاترین سطح سرمایه‌گذاری در طلای فیزیکی را نمایندگی می‌کند. "
                "این وزن با کمترین اجرت ساخت نسبت به گرم، بیشترین بازدهی را برای سرمایه‌گذاران فراهم می‌سازد.\n\n"
                "حداقل اجرت: پایین‌ترین هزینه ساخت به ازای هر گرم طلا.\n"
                "حفظ ثروت: ذخیره ایمن سرمایه‌های بسیار بزرگ در قالبی فشرده.\n"
                "اعتبار و اطمینان: انتخاب نهادهای مالی و سرمایه‌گذاران نهادی.\n\n"
                "با خلوص ۹۹۵ (۲۴ عیار)، هر شمش دارای شناسه یکتا، گواهی اصالت و بسته‌بندی ایمن ویژه است."
            ),
        }

        silver_descriptions = {
            0.1: (
                "شمش نقره ۱۰۰ سوتی؛ آغاز سرمایه‌گذاری در فلز گران‌بها\n\n"
                "شمش نقره ۱۰۰ سوتی کوچک‌ترین واحد سرمایه‌گذاری در نقره خالص است. "
                "نقره به عنوان دومین فلز گران‌بهای جهان، فرصتی منحصربه‌فرد برای تنوع‌بخشی به سبد سرمایه فراهم می‌کند.\n\n"
                "شروع آسان: ورود به بازار نقره با کمترین سرمایه اولیه.\n"
                "هدیه‌ای خاص: گزینه‌ای متفاوت و ارزشمند برای مناسبت‌ها.\n"
                "پتانسیل رشد: نقره با کاربردهای صنعتی گسترده، پتانسیل رشد بالایی دارد.\n\n"
                "هر قطعه با خلوص ۹۹۹.۹ و کد رهگیری اختصاصی عرضه می‌شود."
            ),
            0.2: (
                "شمش نقره ۲۰۰ سوتی؛ پس‌انداز هوشمند در نقره\n\n"
                "شمش نقره ۲۰۰ سوتی فرصتی عالی برای پس‌انداز تدریجی در فلزات گران‌بهاست. "
                "نقره خالص با کاربردهای گسترده صنعتی و سرمایه‌ای، همواره جایگاه ویژه‌ای در سبد دارایی دارد.\n\n"
                "سرمایه‌گذاری تدریجی: خرید منظم و ساخت ذخیره نقره.\n"
                "تنوع سبد: مکمل مناسب برای سرمایه‌گذاری طلا.\n"
                "هدیه‌ای ماندگار: زیبا و ارزشمند برای هر مناسبتی.\n\n"
                "با خلوص ۹۹۹.۹ و بسته‌بندی امن، اصالت هر شمش تضمین شده است."
            ),
            0.5: (
                "شمش نقره نیم گرمی؛ ارزش واقعی در وزنی سبک\n\n"
                "شمش نقره ۵۰۰ سوتی (نیم گرم) یکی از پرطرفدارترین وزن‌ها برای سرمایه‌گذاری خُرد در نقره است. "
                "نقره خالص به عنوان فلزی با ارزش ذاتی و کاربرد صنعتی، آینده‌ای روشن دارد.\n\n"
                "سرمایه‌گذاری هدفمند: قدمی معنادار در مسیر ذخیره نقره.\n"
                "هدیه برازنده: مناسب عروسی، نامزدی و جشن‌های خانوادگی.\n"
                "نقدشوندگی: فروش آسان با قیمت روز بازار.\n\n"
                "هر قطعه با خلوص ۹۹۹.۹، کد رهگیری و بسته‌بندی ایمن ارائه می‌شود."
            ),
            1.0: (
                "شمش نقره یک گرمی؛ استاندارد سرمایه‌گذاری خُرد\n\n"
                "شمش نقره یک گرمی واحد استاندارد سرمایه‌گذاری در نقره خالص است. "
                "نقره با نقش دوگانه خود به‌عنوان فلز گران‌بها و فلز صنعتی، یکی از هوشمندانه‌ترین "
                "گزینه‌های سرمایه‌گذاری محسوب می‌شود.\n\n"
                "ارزش پایدار: نقره همواره به عنوان ذخیره ارزش شناخته شده است.\n"
                "کاربرد صنعتی: تقاضای رو به رشد در صنایع الکترونیک و انرژی خورشیدی.\n"
                "هدیه‌ای بی‌نظیر: زیبا، ماندگار و دارای ارزش واقعی.\n\n"
                "با خلوص ۹۹۹.۹ و گواهی اصالت، هر شمش تحت نظارت دقیق تولید می‌شود."
            ),
            2.5: (
                "شمش نقره ۲.۵ گرمی؛ انتخاب هوشمندانه تنوع‌بخشی\n\n"
                "شمش نقره ۲.۵ گرمی فرصتی عالی برای تنوع‌بخشی به سبد فلزات گران‌بهاست. "
                "ترکیب نقره با طلا در سبد سرمایه، ریسک را کاهش و پتانسیل سود را افزایش می‌دهد.\n\n"
                "تنوع سبد: مکمل ایده‌آل سرمایه‌گذاری طلا.\n"
                "پتانسیل رشد: نقره نسبت به طلا پتانسیل رشد درصدی بالاتری دارد.\n"
                "هدیه‌ای لوکس: انتخابی متفاوت و شایسته برای مناسبت‌های ویژه.\n\n"
                "هر شمش با خلوص ۹۹۹.۹، شناسه یکتا و بسته‌بندی امن عرضه می‌شود."
            ),
            5.0: (
                "شمش نقره ۵ گرمی؛ ذخیره ارزش پایدار\n\n"
                "شمش نقره ۵ گرمی انتخاب محبوب سرمایه‌گذارانی است که به ارزش بلندمدت نقره ایمان دارند. "
                "رشد تقاضای صنعتی نقره در فناوری‌های نوین، آینده‌ای درخشان برای این فلز ترسیم کرده است.\n\n"
                "سرمایه‌گذاری جدی: حجم مناسب برای شروع ذخیره‌سازی نقره.\n"
                "رشد صنعتی: افزایش تقاضا در صنایع فناوری و انرژی سبز.\n"
                "نقدشوندگی: فروش آسان و سریع در بازار.\n\n"
                "تولید شده با خلوص ۹۹۹.۹ و دارای کد رهگیری و گواهی اصالت."
            ),
            10.0: (
                "شمش نقره ۱۰ گرمی؛ سرمایه‌گذاری حرفه‌ای در نقره\n\n"
                "شمش نقره ۱۰ گرمی وزنی استاندارد در بازار نقره است و برای سرمایه‌گذاران جدی طراحی شده. "
                "نقره خالص در کنار طلا، ستون دوم سبد فلزات گران‌بها محسوب می‌شود.\n\n"
                "وزن استاندارد: شناخته‌شده و قابل معامله در بازارهای معتبر.\n"
                "حفظ ارزش: محافظت از سرمایه در برابر تورم.\n"
                "بازدهی بالقوه: پتانسیل رشد قیمت نقره در بلندمدت.\n\n"
                "با خلوص ۹۹۹.۹، هر شمش دارای شناسه یکتا و گواهی اصالت است."
            ),
            20.0: (
                "شمش نقره ۲۰ گرمی؛ گنجینه نقره‌ای\n\n"
                "شمش نقره ۲۰ گرمی ذخیره‌ای ارزشمند برای سرمایه‌گذاران بلندمدت است. "
                "با افزایش کاربردهای صنعتی نقره، این فلز همچنان یکی از جذاب‌ترین گزینه‌های سرمایه‌گذاری باقی مانده است.\n\n"
                "سرمایه‌گذاری بلندمدت: حفظ ثروت در قالب فلز گران‌بها.\n"
                "تقاضای رو به رشد: افزایش مصرف صنعتی نقره در جهان.\n"
                "نقدشوندگی تضمینی: فروش آسان با قیمت بازار.\n\n"
                "هر قطعه با خلوص ۹۹۹.۹ و بسته‌بندی ایمن و کد رهگیری عرضه می‌شود."
            ),
            31.1: (
                "شمش نقره یک اونسی؛ استاندارد بین‌المللی نقره\n\n"
                "شمش نقره یک اونس (۳۱.۱ گرم) واحد اصلی قیمت‌گذاری نقره در بازارهای جهانی است. "
                "خرید نقره اونسی، سرمایه‌گذاری مطابق با استانداردهای بین‌المللی است.\n\n"
                "مرجع قیمت: قیمت نقره در بازارهای جهانی بر اساس اونس تعیین می‌شود.\n"
                "سرمایه‌گذاری حرفه‌ای: انتخاب سرمایه‌گذاران باتجربه و نهادهای مالی.\n"
                "نقدشوندگی بالا: بالاترین سطح تقاضا و سهولت معامله.\n\n"
                "با خلوص ۹۹۹.۹ و گواهی اصالت، هر شمش اونسی با شناسه یکتا عرضه می‌شود."
            ),
            50.0: (
                "شمش نقره ۵۰ گرمی؛ ذخیره سنگین فلز گران‌بها\n\n"
                "شمش نقره ۵۰ گرمی برای سرمایه‌گذارانی طراحی شده که حجم بالایی از نقره خالص را ذخیره می‌کنند. "
                "این وزن با اجرت ساخت بهینه‌تر، بازدهی بهتری نسبت به وزن‌های سبک‌تر ارائه می‌دهد.\n\n"
                "بهینه‌سازی هزینه: اجرت ساخت کمتر به ازای هر گرم.\n"
                "سرمایه‌گذاری کلان: ذخیره حجم بالای نقره خالص.\n"
                "ارزش صنعتی: نقره با تقاضای رو به رشد صنعتی، آینده‌ای روشن دارد.\n\n"
                "تولید شده با خلوص ۹۹۹.۹، دارای کد رهگیری و بسته‌بندی ایمن ویژه."
            ),
            100.0: (
                "شمش نقره ۱۰۰ گرمی؛ اوج سرمایه‌گذاری نقره‌ای\n\n"
                "شمش نقره ۱۰۰ گرمی بالاترین سطح سرمایه‌گذاری در نقره فیزیکی را نمایندگی می‌کند. "
                "با کمترین اجرت ساخت و بالاترین کارایی، این وزن انتخاب سرمایه‌گذاران نهادی است.\n\n"
                "حداقل اجرت: پایین‌ترین هزینه ساخت به ازای هر گرم نقره.\n"
                "حفظ ثروت: ذخیره امن سرمایه‌های بزرگ.\n"
                "تقاضای جهانی: نقره همواره جزو پرتقاضاترین فلزات گران‌بها بوده است.\n\n"
                "با خلوص ۹۹۹.۹، هر شمش دارای شناسه یکتا، گواهی اصالت و بسته‌بندی ایمن ویژه است."
            ),
        }

        product_types = [
            {
                "folder": "شمش طلا با بسته بندی",
                "excel_sheet": "شمش طلاملا",
                "slug": "gold-talamala",
                "category_slug": "gold-talamala",
                "name_prefix": "شمش طلا طلاملا",
                "purity": 995,
                "metal": "gold",
            },
            {
                "folder": "شمش طلا بدون بسته بندی",
                "excel_sheet": "شمش سرمایه ای",
                "slug": "gold-investment",
                "category_slug": "gold-investment",
                "name_prefix": "شمش طلا سرمایه‌ای",
                "purity": 995,
                "metal": "gold",
            },
            {
                "folder": "شمش نقره با بسته بندی",
                "excel_sheet": "شمش نقره طلاملا",
                "slug": "silver-talamala",
                "category_slug": "silver-talamala",
                "name_prefix": "شمش نقره طلاملا",
                "purity": 999.9,
                "metal": "silver",
            },
            {
                "folder": "شمش نقره بدون بسته بندی",
                "excel_sheet": "شمش سرمایه ای نقره",
                "slug": "silver-investment",
                "category_slug": "silver-investment",
                "name_prefix": "شمش نقره سرمایه‌ای",
                "purity": 999.9,
                "metal": "silver",
            },
        ]

        weights_def = [
            (0.1,  "۱۰۰ سوت",  {"شمش نقره با بسته بندی": "0.1 g"}),
            (0.2,  "۲۰۰ سوت",  {}),
            (0.5,  "۵۰۰ سوت",  {}),
            (1.0,  "۱ گرم",     {}),
            (2.5,  "۲.۵ گرم",   {}),
            (5.0,  "۵ گرم",     {}),
            (10.0, "۱۰ گرم",    {}),
            (20.0, "۲۰ گرم",    {}),
            (31.1, "۱ اونس",    {
                "شمش طلا با بسته بندی": "1ounce",
                "شمش طلا بدون بسته بندی": "1ounce",
                "شمش نقره با بسته بندی": "1onuce",
                "شمش نقره بدون بسته بندی": "1onuce",
            }),
            (50.0, "۵۰ گرم",    {"شمش طلا بدون بسته بندی": "50G"}),
            (100.0,"۱۰۰ گرم",   {}),
        ]

        def weight_to_folder(weight_grams, ptype_folder, overrides):
            if ptype_folder in overrides:
                return overrides[ptype_folder]
            if weight_grams == 31.1:
                return "1ounce"
            if weight_grams == int(weight_grams):
                return f"{int(weight_grams)}g"
            return f"{weight_grams}g"

        # Read wage data from Excel
        EXCEL_PATH = os.path.join(IMG_SRC_BASE, "سطوح قیمت گذاری طلا ونقره.xlsx")
        wage_data = {}

        if os.path.exists(EXCEL_PATH):
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
            for ptype in product_types:
                sheet_name = ptype["excel_sheet"]
                if sheet_name not in wb.sheetnames:
                    print(f"  ! Excel sheet '{sheet_name}' not found")
                    continue
                ws = wb[sheet_name]
                slug_wages = {}
                for row in ws.iter_rows(min_row=3, max_col=5, values_only=True):
                    weight_val, t1, t2, t3, t4 = row
                    if weight_val is None or t4 is None:
                        continue
                    w = round(float(weight_val), 2)
                    slug_wages[w] = [
                        round(float(t1 or 0) * 100, 2),
                        round(float(t2 or 0) * 100, 2),
                        round(float(t3 or 0) * 100, 2),
                        round(float(t4 or 0) * 100, 2),
                    ]
                wage_data[ptype["slug"]] = slug_wages
            wb.close()
            print(f"  Wage data loaded from Excel ({len(wage_data)} types)")
        else:
            print(f"  ! Excel file not found: {EXCEL_PATH}")

        print(f"\n[5] Products ({len(weights_def)} weights x {len(product_types)} types)")

        product_map = {}
        default_design = db.query(CardDesign).first()
        default_package = db.query(PackageType).first()

        for ptype in product_types:
            cat = cat_map.get(ptype["category_slug"])
            type_wages = wage_data.get(ptype["slug"], {})
            desc_map = gold_descriptions if ptype["metal"] == "gold" else silver_descriptions

            for weight_grams, weight_label, folder_overrides in weights_def:
                name = f"{ptype['name_prefix']} {weight_label}"
                tier_wages_list = type_wages.get(weight_grams, [0, 0, 0, 0])
                end_customer_wage = tier_wages_list[3]
                description = desc_map.get(weight_grams)

                existing = db.query(Product).filter(Product.name == name).first()
                if existing:
                    p = existing
                    product_map[name] = p
                    if end_customer_wage and p.wage != end_customer_wage:
                        p.wage = end_customer_wage
                    expected_metal = ptype.get("metal", "gold")
                    if p.metal_type != expected_metal:
                        p.metal_type = expected_metal
                    if description:
                        p.description = description

                    # Refresh images
                    old_images = db.query(ProductImage).filter(ProductImage.product_id == p.id).all()
                    for old_img in old_images:
                        old_file = os.path.join(PROJECT_ROOT, old_img.file_path)
                        if os.path.exists(old_file):
                            os.remove(old_file)
                        db.delete(old_img)
                    db.flush()
                else:
                    p = Product(
                        name=name,
                        weight=weight_grams,
                        purity=ptype["purity"],
                        metal_type=ptype.get("metal", "gold"),
                        card_design_id=default_design.id if default_design else None,
                        package_type_id=default_package.id if default_package else None,
                        wage=end_customer_wage,
                        is_wage_percent=True,
                        is_active=True,
                        description=description,
                    )
                    db.add(p)
                    db.flush()
                    if cat:
                        db.add(ProductCategoryLink(product_id=p.id, category_id=cat.id))
                    print(f"  + {name} ({weight_grams}g, wage={end_customer_wage}%)")

                # Copy images from _private
                folder_name = weight_to_folder(weight_grams, ptype["folder"], folder_overrides)
                img_src_dir = os.path.join(IMG_SRC_BASE, ptype["folder"], folder_name)
                if os.path.isdir(img_src_dir):
                    src_files = []
                    for root, _dirs, files in os.walk(img_src_dir):
                        for f in sorted(files):
                            if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                                src_files.append(os.path.join(root, f))
                    src_files.sort()
                    # Only keep first 2 images (front + back)
                    src_files = src_files[:2]
                    for idx, src_path in enumerate(src_files):
                        ext = os.path.splitext(src_path)[1].lower()
                        clean_name = f"{ptype['slug']}_{folder_name}_{idx + 1}{ext}"
                        dst_path = os.path.join(IMG_DST_BASE, clean_name)
                        shutil.copy2(src_path, dst_path)
                        rel_path = f"static/uploads/products/{clean_name}"
                        db.add(ProductImage(product_id=p.id, file_path=rel_path, is_default=(idx == 0)))

                product_map[name] = p

        db.flush()
        print(f"  {len(product_map)} products ready")

        # ==========================================
        # 6. Geographic Data (Iran provinces/cities)
        # ==========================================
        print("\n[6] Geographic Data")

        geo_data = {
            "تهران": {
                "cities": ["تهران", "شهریار", "اسلامشهر", "ورامین", "ری", "قدس", "پاکدشت", "ملارد", "رباط‌کریم", "بومهن", "دماوند"],
                "districts": {
                    "تهران": ["همه محله‌ها", "تهرانپارس", "نارمک", "پیروزی", "ونک", "سعادت‌آباد", "شهرک غرب",
                              "پونک", "ستارخان", "آریاشهر", "جنت‌آباد", "اکباتان", "تجریش", "ولیعصر",
                              "میرداماد", "یوسف‌آباد", "امیرآباد", "انقلاب", "بازار", "مولوی", "جمهوری",
                              "فردوسی", "آزادی", "صادقیه", "المهدی", "چیتگر", "اندرزگو", "زعفرانیه"],
                },
            },
            "اصفهان": {
                "cities": ["اصفهان", "کاشان", "نجف‌آباد", "خمینی‌شهر", "شاهین‌شهر", "فلاورجان", "لنجان"],
                "districts": {"اصفهان": ["همه محله‌ها", "چهارباغ", "جلفا", "خوراسگان", "شاهین‌ویلا", "ملک‌شهر"]},
            },
            "فارس": {
                "cities": ["شیراز", "مرودشت", "کازرون", "فسا", "جهرم", "لار", "داراب", "آباده"],
                "districts": {"شیراز": ["همه محله‌ها", "زند", "معالی‌آباد", "قصرالدشت", "ملاصدرا", "صدرا"]},
            },
            "خراسان رضوی": {
                "cities": ["مشهد", "نیشابور", "سبزوار", "تربت‌حیدریه", "قوچان", "کاشمر", "گناباد", "چناران"],
                "districts": {"مشهد": ["همه محله‌ها", "وکیل‌آباد", "احمدآباد", "هاشمیه", "الهیه", "سجاد"]},
            },
            "آذربایجان شرقی": {
                "cities": ["تبریز", "مراغه", "مرند", "میانه", "اهر", "سراب", "بناب", "شبستر"],
                "districts": {"تبریز": ["همه محله‌ها", "ولیعصر", "آزادی", "ائل‌گلی", "باغمیشه", "رشدیه"]},
            },
            "خوزستان": {
                "cities": ["اهواز", "آبادان", "خرمشهر", "دزفول", "شوشتر", "ماهشهر", "بهبهان", "اندیمشک"],
                "districts": {"اهواز": ["همه محله‌ها", "کیانپارس", "گلستان", "زیتون", "پادادشهر"]},
            },
            "گیلان": {
                "cities": ["رشت", "بندر انزلی", "لاهیجان", "لنگرود", "آستارا", "تالش", "فومن", "صومعه‌سرا"],
                "districts": {"رشت": ["همه محله‌ها"]},
            },
            "مازندران": {
                "cities": ["ساری", "بابل", "آمل", "قائم‌شهر", "بهشهر", "تنکابن", "رامسر", "نوشهر", "چالوس", "نور"],
                "districts": {"ساری": ["همه محله‌ها"]},
            },
            "کرمان": {
                "cities": ["کرمان", "رفسنجان", "جیرفت", "سیرجان", "بم", "زرند", "بردسیر"],
                "districts": {"کرمان": ["همه محله‌ها"]},
            },
            "البرز": {
                "cities": ["کرج", "فردیس", "نظرآباد", "هشتگرد", "محمدشهر", "مشکین‌دشت"],
                "districts": {"کرج": ["همه محله‌ها", "گوهردشت", "مهرشهر", "عظیمیه", "گلشهر", "جهانشهر"]},
            },
            "قم": {"cities": ["قم"], "districts": {"قم": ["همه محله‌ها"]}},
            "کرمانشاه": {
                "cities": ["کرمانشاه", "اسلام‌آباد غرب", "سنقر", "کنگاور", "پاوه", "جوانرود", "هرسین"],
                "districts": {"کرمانشاه": ["همه محله‌ها"]},
            },
            "آذربایجان غربی": {
                "cities": ["ارومیه", "خوی", "مهاباد", "بوکان", "میاندوآب", "سلماس", "پیرانشهر", "نقده"],
                "districts": {"ارومیه": ["همه محله‌ها"]},
            },
            "هرمزگان": {
                "cities": ["بندرعباس", "قشم", "کیش", "میناب", "بندر لنگه", "حاجی‌آباد"],
                "districts": {"بندرعباس": ["همه محله‌ها"]},
            },
            "سیستان و بلوچستان": {
                "cities": ["زاهدان", "چابهار", "ایرانشهر", "سراوان", "زابل", "خاش", "نیک‌شهر"],
                "districts": {"زاهدان": ["همه محله‌ها"]},
            },
            "لرستان": {
                "cities": ["خرم‌آباد", "بروجرد", "دورود", "الیگودرز", "کوهدشت", "نورآباد"],
                "districts": {"خرم‌آباد": ["همه محله‌ها"]},
            },
            "همدان": {
                "cities": ["همدان", "ملایر", "نهاوند", "تویسرکان", "بهار", "اسدآباد", "کبودرآهنگ"],
                "districts": {"همدان": ["همه محله‌ها"]},
            },
            "مرکزی": {
                "cities": ["اراک", "ساوه", "خمین", "دلیجان", "محلات", "شازند", "تفرش"],
                "districts": {"اراک": ["همه محله‌ها"]},
            },
            "گلستان": {
                "cities": ["گرگان", "گنبد کاووس", "علی‌آباد کتول", "بندر ترکمن", "آق‌قلا", "مینودشت"],
                "districts": {"گرگان": ["همه محله‌ها"]},
            },
            "اردبیل": {
                "cities": ["اردبیل", "پارس‌آباد", "خلخال", "مشگین‌شهر", "نمین", "نیر"],
                "districts": {"اردبیل": ["همه محله‌ها"]},
            },
            "یزد": {
                "cities": ["یزد", "میبد", "اردکان", "تفت", "بافق", "ابرکوه", "مهریز"],
                "districts": {"یزد": ["همه محله‌ها"]},
            },
            "زنجان": {
                "cities": ["زنجان", "ابهر", "خدابنده", "خرمدره", "ماهنشان", "طارم"],
                "districts": {"زنجان": ["همه محله‌ها"]},
            },
            "سمنان": {
                "cities": ["سمنان", "شاهرود", "دامغان", "گرمسار", "مهدی‌شهر"],
                "districts": {"سمنان": ["همه محله‌ها"]},
            },
            "قزوین": {
                "cities": ["قزوین", "تاکستان", "بویین‌زهرا", "آبیک", "آوج", "الموت"],
                "districts": {"قزوین": ["همه محله‌ها"]},
            },
            "بوشهر": {
                "cities": ["بوشهر", "برازجان", "کنگان", "گناوه", "دیلم", "دیر", "جم"],
                "districts": {"بوشهر": ["همه محله‌ها"]},
            },
            "کهگیلویه و بویراحمد": {
                "cities": ["یاسوج", "دهدشت", "گچساران", "دوگنبدان", "لیکک"],
                "districts": {"یاسوج": ["همه محله‌ها"]},
            },
            "چهارمحال و بختیاری": {
                "cities": ["شهرکرد", "بروجن", "فارسان", "لردگان", "اردل", "کیار"],
                "districts": {"شهرکرد": ["همه محله‌ها"]},
            },
            "کردستان": {
                "cities": ["سنندج", "سقز", "مریوان", "بانه", "قروه", "بیجار", "دیواندره", "کامیاران"],
                "districts": {"سنندج": ["همه محله‌ها"]},
            },
            "ایلام": {
                "cities": ["ایلام", "دهلران", "ایوان", "آبدانان", "مهران", "دره‌شهر"],
                "districts": {"ایلام": ["همه محله‌ها"]},
            },
            "خراسان شمالی": {
                "cities": ["بجنورد", "شیروان", "اسفراین", "جاجرم", "فاروج"],
                "districts": {"بجنورد": ["همه محله‌ها"]},
            },
            "خراسان جنوبی": {
                "cities": ["بیرجند", "قاین", "طبس", "فردوس", "نهبندان", "سربیشه"],
                "districts": {"بیرجند": ["همه محله‌ها"]},
            },
        }

        existing_prov = db.query(GeoProvince).count()
        if existing_prov == 0:
            sort_idx = 0
            for prov_name, prov_info in geo_data.items():
                sort_idx += 1
                prov = GeoProvince(name=prov_name, sort_order=sort_idx)
                db.add(prov)
                db.flush()
                for city_name in prov_info["cities"]:
                    city = GeoCity(province_id=prov.id, name=city_name)
                    db.add(city)
                    db.flush()
                    district_list = prov_info.get("districts", {}).get(city_name, [])
                    for dist_name in district_list:
                        db.add(GeoDistrict(city_id=city.id, name=dist_name))
            db.flush()
            total_cities = db.query(GeoCity).count()
            total_districts = db.query(GeoDistrict).count()
            print(f"  + {len(geo_data)} provinces, {total_cities} cities, {total_districts} districts")
        else:
            print(f"  = geo data exists ({existing_prov} provinces)")

        # ==========================================
        # 7. Dealer Tiers
        # ==========================================
        print("\n[7] Dealer Tiers")

        tiers_data = [
            {"name": "پخش", "slug": "distributor", "sort_order": 1, "is_end_customer": False},
            {"name": "بنکدار", "slug": "wholesaler", "sort_order": 2, "is_end_customer": False},
            {"name": "فروشگاه", "slug": "store", "sort_order": 3, "is_end_customer": False},
            {"name": "مشتری نهایی", "slug": "end_customer", "sort_order": 4, "is_end_customer": True},
        ]

        tier_map = {}
        for td in tiers_data:
            existing = db.query(DealerTier).filter(DealerTier.slug == td["slug"]).first()
            if not existing:
                tier = DealerTier(**td, is_active=True)
                db.add(tier)
                db.flush()
                tier_map[td["slug"]] = tier
                print(f"  + {td['name']} ({td['slug']})")
            else:
                tier_map[td["slug"]] = existing
                print(f"  = exists: {td['name']}")

        db.flush()

        # ==========================================
        # 7.5 Product Tier Wages (from Excel)
        # ==========================================
        print("\n[7.5] Product Tier Wages")

        tier_slugs_order = ["distributor", "wholesaler", "store", "end_customer"]

        existing_tw = db.query(ProductTierWage).count()
        if existing_tw > 0:
            db.query(ProductTierWage).delete()
            db.flush()
            print(f"  ~ deleted {existing_tw} old tier wages, refreshing...")

        tw_count = 0
        cat_id_to_slug = {cat_obj.id: slug for slug, cat_obj in cat_map.items()}

        all_products = db.query(Product).filter(Product.is_active == True).all()
        for p in all_products:
            type_slug = None
            for cid in p.category_ids:
                if cid in cat_id_to_slug:
                    type_slug = cat_id_to_slug[cid]
                    break
            if not type_slug or type_slug not in wage_data:
                continue
            weight_key = float(p.weight)
            wages = wage_data[type_slug].get(weight_key)
            if not wages:
                continue
            for idx, tier_slug in enumerate(tier_slugs_order):
                tier = tier_map.get(tier_slug)
                if tier:
                    db.add(ProductTierWage(product_id=p.id, tier_id=tier.id, wage_percent=wages[idx]))
                    tw_count += 1

        db.flush()
        print(f"  + {tw_count} tier wages created")

        # ==========================================
        # 8. Real Bars (from bars_data.json)
        # ==========================================
        print("\n[8] Real Bars")

        bars_json_path = os.path.join(PROJECT_ROOT, "_private", "bars_data.json")
        if not os.path.exists(bars_json_path):
            print(f"  ! bars_data.json not found — skipping bar import")
        else:
            with open(bars_json_path, "r", encoding="utf-8") as f:
                bars_data = json.load(f)

            # Build product name → id map
            v4_name_to_id = {p.name: p.id for p in db.query(Product).all()}

            # Resolve type_slug + weight → product name
            def resolve_product_name(type_slug, weight):
                for ptype in product_types:
                    if ptype["slug"] == type_slug:
                        for w, label, _ in weights_def:
                            if abs(w - weight) < 0.01:
                                return f"{ptype['name_prefix']} {label}"
                return None

            # Get batch
            batch = db.query(Batch).filter(Batch.batch_number == "BATCH-20241203").first()
            batch_id = batch.id if batch else None

            # Get existing serials to skip duplicates
            existing_serials = {row[0] for row in db.query(Bar.serial_code).all()}

            # Create customers for sold bars
            customer_mobile_to_id = {}
            for mobile, cdata in bars_data.get("customers", {}).items():
                existing_c = db.query(User).filter(User.mobile == mobile).first()
                if existing_c:
                    customer_mobile_to_id[mobile] = existing_c.id
                else:
                    c = User(
                        first_name=cdata["first_name"],
                        last_name=cdata["last_name"],
                        national_id=cdata["national_id"],
                        mobile=mobile,
                        is_customer=True,
                        is_active=True,
                    )
                    db.add(c)
                    db.flush()
                    customer_mobile_to_id[mobile] = c.id
                    print(f"  + Customer: {mobile}")

            # Insert assigned bars
            inserted = 0
            for key, serials in bars_data.get("assigned", {}).items():
                parts = key.split("|")
                type_slug, weight = parts[0], float(parts[1])
                pname = resolve_product_name(type_slug, weight)
                if not pname or pname not in v4_name_to_id:
                    print(f"  ! Product not found: {key}")
                    continue
                pid = v4_name_to_id[pname]
                for serial in serials:
                    if serial in existing_serials:
                        continue
                    db.add(Bar(serial_code=serial, status="Sold", product_id=pid, batch_id=batch_id,
                              delivered_at=datetime.now(timezone.utc)))
                    existing_serials.add(serial)
                    inserted += 1

            # Insert sold bars
            for sb in bars_data.get("sold", []):
                serial = sb["serial"]
                if serial in existing_serials:
                    continue
                pname = resolve_product_name(sb["type"], sb["weight"])
                if not pname or pname not in v4_name_to_id:
                    continue
                cid = customer_mobile_to_id.get(sb.get("customer_mobile"))
                db.add(Bar(
                    serial_code=serial,
                    status="Sold",
                    product_id=v4_name_to_id[pname],
                    batch_id=batch_id,
                    customer_id=cid,
                    delivered_at=datetime.now(timezone.utc),
                ))
                existing_serials.add(serial)
                inserted += 1

            db.flush()
            print(f"  + {inserted} bars imported")
            print(f"  + Customers: {len(customer_mobile_to_id)}")

        # ==========================================
        # 9. New Inventory Bars (gold only)
        # ==========================================
        print("\n[9] New Inventory Bars (gold only)")

        # Collect all existing serials (old + any already in DB)
        all_serials = {row[0] for row in db.query(Bar.serial_code).all()}

        def gen_serial(existing: set) -> str:
            chars = string.ascii_uppercase + string.digits
            while True:
                code = "".join(random.choices(chars, k=8))
                if code not in existing:
                    existing.add(code)
                    return code

        new_bar_counts = {
            0.1: 100,
            0.2: 80,
            0.5: 60,
            1.0: 50,
            2.5: 30,
            5.0: 20,
            10.0: 10,
            20.0: 5,
            31.1: 3,
            50.0: 2,
            100.0: 1,
        }

        # Get batch for new bars
        new_batch = db.query(Batch).filter(Batch.batch_number == "BATCH-20241203").first()
        new_batch_id = new_batch.id if new_batch else None

        gold_slugs = ["gold-talamala", "gold-investment"]
        new_bar_total = 0

        for slug in gold_slugs:
            for weight, count in new_bar_counts.items():
                pname = resolve_product_name(slug, weight)
                if not pname or pname not in v4_name_to_id:
                    continue
                pid = v4_name_to_id[pname]
                for _ in range(count):
                    serial = gen_serial(all_serials)
                    db.add(Bar(serial_code=serial, status="Assigned", product_id=pid, batch_id=new_batch_id))
                    new_bar_total += 1

        db.flush()
        print(f"  + {new_bar_total} new bars generated (gold-talamala + gold-investment)")
        for weight, count in new_bar_counts.items():
            print(f"    {weight}g: {count} × 2 = {count * 2}")

        # ==========================================
        # Commit
        # ==========================================
        db.commit()

        print("\n" + "=" * 50)
        print("  Production seed completed!")
        print("=" * 50)

        print("\n--- Summary ---")
        print(f"  Admin users:    {db.query(User).filter(User.is_admin == True).count()}")
        print(f"  Customers:      {db.query(User).filter(User.is_customer == True).count()}")
        print(f"  Settings:       {db.query(SystemSetting).count()}")
        print(f"  Products:       {db.query(Product).count()}")
        print(f"  Categories:     {db.query(ProductCategory).count()}")
        print(f"  Provinces:      {db.query(GeoProvince).count()}")
        print(f"  Cities:         {db.query(GeoCity).count()}")
        print(f"  Card Designs:   {db.query(CardDesign).count()}")
        print(f"  Package Types:  {db.query(PackageType).count()}")
        print(f"  Batches:        {db.query(Batch).count()}")
        print(f"  Bars:           {db.query(Bar).count()}")
        print(f"  Dealer Tiers:   {db.query(DealerTier).count()}")
        print(f"  Tier Wages:     {db.query(ProductTierWage).count()}")
        print(f"  Dealers:        {db.query(User).filter(User.is_dealer == True).count()}")

        print(f"\n--- Admins ---")
        for au in db.query(User).filter(User.is_admin == True).all():
            print(f"  {au.admin_role}: {au.mobile}")

    except Exception as e:
        db.rollback()
        print(f"\nSeed failed: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


def reset_and_seed():
    from sqlalchemy import text
    with engine.connect() as conn:
        conn.execute(text("DROP TABLE IF EXISTS location_transfers CASCADE"))
        conn.execute(text("DROP TABLE IF EXISTS package_images CASCADE"))
        conn.execute(text("DROP TABLE IF EXISTS design_images CASCADE"))
        # Drop legacy tables from pre-unified-user era
        conn.execute(text("DROP TABLE IF EXISTS customers CASCADE"))
        conn.execute(text("DROP TABLE IF EXISTS dealers CASCADE"))
        conn.execute(text("DROP TABLE IF EXISTS system_users CASCADE"))
        conn.commit()
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    print("All tables recreated")
    seed()


if __name__ == "__main__":
    if "--reset" in sys.argv:
        confirm = input("This will DELETE ALL DATA. Type 'yes' to confirm: ")
        if confirm.strip().lower() == "yes":
            reset_and_seed()
        else:
            print("Aborted.")
    else:
        seed()
