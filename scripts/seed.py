"""
TalaMala v4 - Comprehensive Database Seeder
==============================================
Seeds ALL modules with initial data for testing.

Usage:
    python scripts/seed.py          # Full seed
    python scripts/seed.py --reset  # Drop all data and reseed

Modules seeded:
  1. Admin users (admin + operator)
  2. Test customers
  3. System settings (gold price, tax, shipping, etc.)
  4. Products (TalaMala v2 gold bar weights)
  5. Card designs + Package types + Batches
  6. Locations (branches + postal hub)
  7. Bars (assigned, with location)
  8. Sample coupons (public + mobile-restricted)
  9. Test wallet credit
"""

import sys
import os
import io
import random
import string
import shutil
from datetime import datetime, timezone
import openpyxl

# Fix Windows console encoding for Persian text
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config.database import SessionLocal, Base, engine
from modules.user.models import User
from modules.admin.models import SystemSetting
from modules.customer.address_models import GeoProvince, GeoCity, GeoDistrict, CustomerAddress
from modules.catalog.models import (
    ProductCategory, ProductCategoryLink, Product, ProductImage, CardDesign, CardDesignImage,
    PackageType, PackageTypeImage, Batch, BatchImage, ProductTierWage,
)
from modules.inventory.models import (
    Bar, BarImage, OwnershipHistory, BarStatus,
    DealerTransfer, BarTransfer,
)
from modules.cart.models import Cart, CartItem
from modules.order.models import Order, OrderItem, OrderStatusLog, OrderStatus, DeliveryMethod, DeliveryStatus
from modules.wallet.models import Account, LedgerEntry, WalletTopup, WithdrawalRequest
from modules.coupon.models import Coupon, CouponMobile, CouponUsage, CouponCategory
from modules.dealer.models import DealerTier, DealerSale, BuybackRequest
from modules.ticket.models import Ticket, TicketMessage, TicketAttachment, TicketStatus, TicketPriority, TicketCategory, SenderType
from modules.review.models import Review, ReviewImage, ProductComment, CommentImage, CommentLike
from modules.dealer_request.models import DealerRequest, DealerRequestAttachment
from modules.pricing.models import Asset, GOLD_18K, SILVER


def generate_serial() -> str:
    chars = string.ascii_uppercase + string.digits
    return ''.join(random.choices(chars, k=8))


def ensure_tables():
    """Create all tables if they don't exist (safe to call multiple times)."""
    print("[0/9] Ensuring all tables exist...")
    Base.metadata.create_all(bind=engine)
    print("  + All tables OK\n")


def ensure_schema_updates():
    """Add missing columns to existing tables (create_all won't do this)."""
    from sqlalchemy import text, inspect
    print("[0.5] Checking schema updates...")

    insp = inspect(engine)
    updates = 0

    # --- products: description ---
    if "products" in insp.get_table_names():
        prod_cols = {c["name"] for c in insp.get_columns("products")}
        with engine.begin() as conn:
            if "description" not in prod_cols:
                conn.execute(text("ALTER TABLE products ADD COLUMN description TEXT"))
                print("  + products.description added")
                updates += 1

    if updates == 0:
        print("  = schema up to date")
    print()


def seed():
    db = SessionLocal()
    try:
        print("=" * 50)
        print("  TalaMala v4 — Comprehensive Seeder")
        print("=" * 50)

        # Ensure all tables exist first
        ensure_tables()
        ensure_schema_updates()

        # ==========================================
        # 1. Admin Users
        # ==========================================
        print("\n[1/9] Admin Users")

        admins_data = [
            {"mobile": "09123456789", "first_name": "مدیر", "last_name": "سیستم", "admin_role": "admin"},
            {"mobile": "09121111111", "first_name": "اپراتور", "last_name": "تهران", "admin_role": "operator"},
            {"mobile": "09121023589", "first_name": "ادمین", "last_name": "۱", "admin_role": "admin"},
            {"mobile": "09120725564", "first_name": "ادمین", "last_name": "۲", "admin_role": "admin"},
        ]
        for data in admins_data:
            existing = db.query(User).filter(User.mobile == data["mobile"]).first()
            if not existing:
                user_obj = User(
                    mobile=data["mobile"],
                    first_name=data["first_name"],
                    last_name=data["last_name"],
                    admin_role=data["admin_role"],
                    is_admin=True,
                    is_customer=True,  # admins can also shop
                )
                db.add(user_obj)
                print(f"  + {data['admin_role']}: {data['mobile']} ({data['first_name']} {data['last_name']})")
            else:
                # Ensure admin flags are set
                if not existing.is_admin:
                    existing.is_admin = True
                    existing.admin_role = data["admin_role"]
                user_obj = existing
                print(f"  = exists: {data['mobile']}")

        # Set operator permissions (after flush to ensure objects exist)
        db.flush()
        op_user = db.query(User).filter(User.mobile == "09121111111").first()
        if op_user:
            op_user.permissions = ["dashboard", "orders", "inventory", "tickets", "customers"]
            print(f"  ~ operator permissions set: {op_user.permissions}")

        # ==========================================
        # 2. Test Customers
        # ==========================================
        print("\n[2/9] Test Customers")

        customers_data = [
            {
                "mobile": "09351234567", "national_id": "0012345678",
                "first_name": "علی", "last_name": "رضایی",
                "customer_type": "real",
                "postal_code": "1234567890",
                "address": "تهران، خیابان ولیعصر، پلاک ۱۰۰",
                "phone": "02112345678",
            },
            {
                "mobile": "09359876543", "national_id": "0087654321",
                "first_name": "مریم", "last_name": "احمدی",
                "customer_type": "legal",
                "company_name": "شرکت زرین تجارت",
                "economic_code": "411111111111",
                "postal_code": "9876543210",
                "address": "اصفهان، خیابان چهارباغ، پلاک ۵۰",
                "phone": "03132005678",
            },
            {
                "mobile": "09131112233", "national_id": "1234567890",
                "first_name": "رضا", "last_name": "محمدی",
                "customer_type": "real",
                "postal_code": "1122334455",
                "address": "شیراز، خیابان زند، پلاک ۲۰",
            },
        ]
        for data in customers_data:
            existing = db.query(User).filter(User.mobile == data["mobile"]).first()
            if not existing:
                db.add(User(**data, is_customer=True))
                print(f"  + {data['first_name']} {data['last_name']}: {data['mobile']}")
            else:
                if not existing.is_customer:
                    existing.is_customer = True
                print(f"  = exists: {data['mobile']}")

        db.flush()

        # ==========================================
        # 3. System Settings
        # ==========================================
        print("\n[3/9] System Settings")

        settings_data = {
            "site_name":            ("طلاملا", "نام سایت"),
            "site_logo":            ("", "لوگوی سایت"),
            "support_phone":        ("02112345678", "شماره پشتیبانی"),
            "support_telegram":     ("@talamala", "تلگرام پشتیبانی"),
            "tax_percent":          ("10", "درصد مالیات بر ارزش افزوده"),
            "min_order_amount":     ("10000000", "حداقل مبلغ سفارش (ریال)"),
            "reservation_minutes":  ("15", "مدت زمان رزرو (دقیقه)"),
            "shipping_cost":        ("500000", "هزینه ارسال پستی (ریال)"),
            "insurance_percent":    ("1.5", "درصد بیمه پست"),
            "insurance_cap":        ("500000000", "سقف بیمه پست (ریال)"),
            "gold_spread_percent":  ("2", "اسپرد تبدیل ریال به طلا (درصد) — deprecated"),
            "gold_fee_customer_percent": ("2", "کارمزد خرید/فروش طلا — مشتری عادی (%)"),
            "gold_fee_dealer_percent":   ("0.5", "کارمزد خرید/فروش طلا — نماینده (%)"),
            "silver_fee_customer_percent": ("1.5", "کارمزد خرید/فروش نقره — مشتری عادی (%)"),
            "silver_fee_dealer_percent":   ("0.3", "کارمزد خرید/فروش نقره — نماینده (%)"),
            "enabled_gateways":     ("sepehr,top,parsian", "درگاه‌های فعال پرداخت (comma-separated)"),
            "shahkar_enabled":      ("false", "فعال/غیرفعال بودن احراز هویت شاهکار"),
            "rasis_pos_enabled":    ("false", "فعال/غیرفعال بودن همگام‌سازی با دستگاه پوز راسیس"),
        }

        for key, (value, desc) in settings_data.items():
            existing = db.query(SystemSetting).filter(SystemSetting.key == key).first()
            if not existing:
                db.add(SystemSetting(key=key, value=value, description=desc))
                print(f"  + {key} = {value}")
            else:
                print(f"  = exists: {key}")

        # Asset prices (gold + silver)
        asset_seed = [
            {
                "asset_code": GOLD_18K,
                "asset_label": "طلای ۱۸ عیار",
                "price_per_gram": 52_000_000,
                "stale_after_minutes": 15,
                "auto_update": True,
                "update_interval_minutes": 5,
                "source_url": "https://goldis.ir/price/api/v1/price/assets/gold18k/final-prices",
            },
            {
                "asset_code": SILVER,
                "asset_label": "نقره خالص",
                "price_per_gram": 550_000,
                "stale_after_minutes": 30,
                "auto_update": False,
                "update_interval_minutes": 30,
                "source_url": None,
            },
        ]
        for data in asset_seed:
            existing = db.query(Asset).filter(Asset.asset_code == data["asset_code"]).first()
            if not existing:
                db.add(Asset(**data))
                print(f"  + Asset: {data['asset_code']} = {data['price_per_gram']:,}")
            else:
                print(f"  = Asset exists: {data['asset_code']}")

        db.flush()

        # ==========================================
        # 4. Card Designs + Package Types + Batches
        # ==========================================
        print("\n[4/9] Catalog Accessories")

        for name in ["طرح کلاسیک", "طرح مدرن", "طرح هدیه", "طرح نوروز"]:
            if not db.query(CardDesign).filter(CardDesign.name == name).first():
                db.add(CardDesign(name=name))
                print(f"  + Design: {name}")

        pkg_data = [
            {"name": "جعبه استاندارد", "price": 0},
            {"name": "جعبه لوکس", "price": 5_000_000},
            {"name": "جعبه هدیه ویژه", "price": 15_000_000},
            {"name": "پاکت ساده", "price": 1_000_000},
        ]
        for pd in pkg_data:
            existing = db.query(PackageType).filter(PackageType.name == pd["name"]).first()
            if not existing:
                db.add(PackageType(name=pd["name"], price=pd["price"], is_active=True))
                print(f"  + Package: {pd['name']} ({pd['price']} rial)")
            else:
                existing.price = pd["price"]
                print(f"  = Package: {pd['name']} (price updated)")

        batches_data = [
            {"batch_number": "B-1403-001", "melt_number": "M-001", "operator": "استاد کریمی", "purity": 750},
            {"batch_number": "B-1403-002", "melt_number": "M-002", "operator": "استاد کریمی", "purity": 750},
            {"batch_number": "B-1403-003", "melt_number": "M-003", "operator": "استاد احمدی", "purity": 750},
        ]
        for bd in batches_data:
            if not db.query(Batch).filter(Batch.batch_number == bd["batch_number"]).first():
                db.add(Batch(**bd))
                print(f"  + Batch: {bd['batch_number']}")

        db.flush()

        # ==========================================
        # 4.5 Product Categories
        # ==========================================
        print("\n[4.5] Product Categories")

        categories_data = [
            ("شمش طلا طلاملا", "gold-talamala", 1),
            ("شمش طلا سرمایه‌ای", "gold-investment", 2),
            ("شمش نقره طلاملا", "silver-talamala", 3),
            ("شمش نقره سرمایه‌ای", "silver-investment", 4),
        ]

        cat_map = {}
        for cat_name, cat_slug, sort_order in categories_data:
            existing = db.query(ProductCategory).filter(ProductCategory.slug == cat_slug).first()
            if not existing:
                cat = ProductCategory(name=cat_name, slug=cat_slug, sort_order=sort_order, is_active=True)
                db.add(cat)
                db.flush()
                cat_map[cat_slug] = cat
                print(f"  + {cat_name}")
            else:
                cat_map[cat_slug] = existing
                print(f"  = exists: {cat_name}")

        db.flush()

        # ==========================================
        # 5. Products (Real products from _private data)
        # ==========================================
        PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        IMG_SRC_BASE = os.path.join(PROJECT_ROOT, "_private", "عکس محصولات")
        IMG_DST_BASE = os.path.join(PROJECT_ROOT, "static", "uploads", "products")
        os.makedirs(IMG_DST_BASE, exist_ok=True)

        # ---- Product descriptions (per weight × metal) ----
        gold_descriptions = {
            0.1: (
                "شمش ۱۰۰ سوتی؛ کوچک‌ترین واحد سرمایه‌گذاری طلا\n\n"
                "شمش ۱۰۰ سوتی (یک‌دهم گرم) نقطه شروعی ایده‌آل برای ورود به بازار طلاست. "
                "این وزن سبک امکان خرید مکرر و پس‌انداز تدریجی را فراهم می‌کند و به‌ویژه "
                "برای هدیه‌های کوچک و نمادین گزینه‌ای بی‌نظیر است.\n\n"
                "هدیه‌ای نمادین: مناسب برای تولد، جشن فارغ‌التحصیلی و مناسبت‌های خاص.\n"
                "پس‌انداز گام‌به‌گام: خرید ماهانه و ساخت سبد طلای شخصی.\n"
                "نقدشوندگی بالا: تبدیل آسان به وجه نقد در هر زمان.\n\n"
                "هر قطعه با خلوص ۹۹۵ (۲۴ عیار) و کد رهگیری اختصاصی عرضه می‌شود "
                "تا اصالت سرمایه شما تضمین گردد."
            ),
            0.2: (
                "شمش ۲۰۰ سوتی؛ قدمی مطمئن در مسیر سرمایه‌گذاری\n\n"
                "شمش ۲۰۰ سوتی (دو‌دهم گرم) تعادلی مناسب بین قیمت پایین و ارزش واقعی طلا "
                "ایجاد می‌کند. این وزن برای کسانی که می‌خواهند به‌صورت منظم طلا پس‌انداز کنند، "
                "انتخابی هوشمندانه است.\n\n"
                "سرمایه‌گذاری تدریجی: امکان خرید مکرر بدون فشار مالی.\n"
                "هدیه‌ای ارزشمند: مناسب برای یادبودها و قدردانی.\n"
                "قابلیت جمع‌آوری: ساخت مجموعه طلای شخصی با خریدهای کوچک.\n\n"
                "با خلوص ۹۹۵ و بسته‌بندی امن، هر شمش دارای شناسه یکتا و گواهی اصالت است."
            ),
            0.5: (
                "شمش نیم گرمی؛ تعادل هوشمندانه ارزش و دسترسی\n\n"
                "شمش ۵۰۰ سوتی (نیم گرم) یکی از محبوب‌ترین وزن‌ها در بازار طلای خرد است. "
                "این وزن، نقطه تلاقی مناسبی بین سرمایه‌گذاری معنادار و قیمت قابل دسترس ایجاد کرده است.\n\n"
                "هدیه‌ای ماندگار: گزینه‌ای برازنده برای عروسی، تولد و مناسبت‌های رسمی.\n"
                "پس‌انداز هدفمند: روشی اصولی برای تبدیل درآمد ماهانه به سرمایه پایدار.\n"
                "نقدشوندگی سریع: فروش آسان در هر زمان با قیمت روز بازار.\n\n"
                "هر قطعه با خلوص ۹۹۵ (۲۴ عیار)، کد رهگیری و بسته‌بندی ایمن عرضه می‌شود."
            ),
            1.0: (
                "شمش یک گرمی؛ جایگاه ویژه در سبد سرمایه‌گذاری\n\n"
                "چرا شمش یک گرمی جایگاه ویژه‌ای در سبد سرمایه‌گذاری دارد؟ وزن یک گرم، "
                "تعادلی هوشمندانه میان نقدشوندگی سریع و قدرت خرید مجدد در دراز مدت ایجاد کرده است.\n\n"
                "هدیه‌ای ماندگار: جایگزینی برازنده برای وجه نقد در مناسبت‌های رسمی و خانوادگی.\n"
                "پس‌انداز مستمر: روشی اصولی برای تبدیل درآمدهای جاری به سرمایه‌ای پایدار.\n"
                "سرمایه‌گذاری امن: مسیری مطمئن برای ورود به بازار طلا، فارغ از ریسک‌های مرسوم.\n\n"
                "خلوص ۹۹۵ (۲۴ عیار): شمش‌های ما با خلوص ۹۹۵ عرضه می‌شوند. "
                "این یعنی شما صاحب خالص‌ترین شکل ممکن طلا هستید.\n"
                "هر قطعه با کد رهگیری اختصاصی و استاندارد تولید، اصالت آن تضمین شده است."
            ),
            2.5: (
                "شمش ۲.۵ گرمی؛ انتخاب حرفه‌ای سرمایه‌گذاران\n\n"
                "شمش ۲.۵ گرمی نقطه عطفی در سرمایه‌گذاری طلا محسوب می‌شود. این وزن، "
                "ضمن حفظ قابلیت نقدشوندگی بالا، ارزش قابل توجهی را در قطعه‌ای فشرده و قابل نگهداری ارائه می‌دهد.\n\n"
                "سرمایه‌گذاری متوسط: مناسب برای کسانی که فراتر از خریدهای خُرد، به دنبال ذخیره ارزش هستند.\n"
                "هدیه‌ای لوکس: انتخابی شایسته برای مناسبت‌های مهم و رویدادهای خاص.\n"
                "پرتفوی متنوع: امکان ترکیب با سایر وزن‌ها برای مدیریت بهتر سبد سرمایه.\n\n"
                "با خلوص ۹۹۵ و گواهی اصالت، هر شمش تحت نظارت دقیق تولید و بسته‌بندی می‌شود."
            ),
            5.0: (
                "شمش ۵ گرمی؛ ذخیره ارزش در ابعاد فشرده\n\n"
                "شمش ۵ گرمی ترکیبی ایده‌آل از ارزش بالا و ابعاد قابل مدیریت است. "
                "این وزن یکی از پرطرفدارترین انتخاب‌ها در میان سرمایه‌گذاران جدی بازار طلاست.\n\n"
                "ذخیره ارزش: نگهداری سرمایه قابل توجه در فضایی بسیار کوچک.\n"
                "نقدشوندگی مناسب: فروش سریع و آسان در بازار با تقاضای بالا.\n"
                "هدیه سرمایه‌ای: انتخابی ممتاز برای هدایای عروسی و رویدادهای مهم زندگی.\n\n"
                "هر شمش با خلوص ۹۹۵ (۲۴ عیار)، کد رهگیری یکتا و بسته‌بندی ایمن ارائه می‌شود."
            ),
            10.0: (
                "شمش ۱۰ گرمی؛ سرمایه‌گذاری قدرتمند و نقدشونده\n\n"
                "شمش ۱۰ گرمی یکی از استانداردترین واحدهای سرمایه‌گذاری در بازار جهانی طلاست. "
                "این وزن، تعادل بهینه‌ای بین حجم سرمایه‌گذاری و سهولت معامله ایجاد می‌کند.\n\n"
                "استاندارد جهانی: وزنی شناخته‌شده و مورد اعتماد در تمام بازارهای طلا.\n"
                "سرمایه‌گذاری کلان: مناسب برای حفظ و رشد سرمایه‌های بزرگ‌تر.\n"
                "نقدشوندگی بالا: قابلیت فروش سریع با کمترین اختلاف قیمت خرید و فروش.\n\n"
                "تولید شده با خلوص ۹۹۵، دارای شناسه یکتا و گواهی اصالت معتبر."
            ),
            20.0: (
                "شمش ۲۰ گرمی؛ گنجینه‌ای برای آینده\n\n"
                "شمش ۲۰ گرمی انتخابی استراتژیک برای سرمایه‌گذاران بلندمدت است. "
                "این وزن ارزش قابل توجهی را در قالبی فشرده و امن ارائه می‌دهد و "
                "گزینه‌ای عالی برای تنوع‌بخشی به سبد دارایی است.\n\n"
                "سرمایه‌گذاری بلندمدت: حفظ ارزش دارایی در برابر تورم و نوسانات اقتصادی.\n"
                "ذخیره امن: نگهداری حجم بالای سرمایه در فضایی بسیار کوچک.\n"
                "نقدشوندگی تضمین‌شده: فروش آسان در بازار با تقاضای همیشگی.\n\n"
                "با خلوص ۹۹۵ (۲۴ عیار) و بسته‌بندی ایمن، اصالت هر قطعه با کد رهگیری تضمین می‌شود."
            ),
            31.1: (
                "شمش یک اونسی؛ استاندارد طلایی بازار جهانی\n\n"
                "شمش یک اونس (۳۱.۱ گرم) معیار اصلی قیمت‌گذاری طلا در بازارهای بین‌المللی است. "
                "خرید شمش اونسی یعنی سرمایه‌گذاری مطابق با استانداردهای جهانی.\n\n"
                "مرجع قیمت جهانی: قیمت طلا در بازارهای بین‌المللی بر اساس اونس تعیین می‌شود.\n"
                "سرمایه‌گذاری حرفه‌ای: انتخاب اول سرمایه‌گذاران حرفه‌ای و نهادهای مالی.\n"
                "نقدشوندگی بی‌رقیب: بالاترین سطح تقاضا و سهولت معامله.\n\n"
                "هر شمش اونسی با خلوص ۹۹۵ (۲۴ عیار)، شناسه یکتا و گواهی اصالت عرضه می‌شود."
            ),
            50.0: (
                "شمش ۵۰ گرمی؛ سرمایه‌گذاری سنگین با حداکثر بازدهی\n\n"
                "شمش ۵۰ گرمی برای سرمایه‌گذارانی طراحی شده که به دنبال ذخیره حجم بالای ارزش هستند. "
                "این وزن، اجرت ساخت کمتری نسبت به وزن‌های سبک‌تر دارد و بازدهی سرمایه‌گذاری را بهبود می‌بخشد.\n\n"
                "بهینه‌سازی هزینه: اجرت ساخت کمتر به نسبت وزن.\n"
                "سرمایه‌گذاری کلان: مناسب برای حفظ ثروت نسل‌ها.\n"
                "نقدشوندگی: قابلیت فروش در بازار حرفه‌ای طلا.\n\n"
                "تولید شده با خلوص ۹۹۵ (۲۴ عیار)، دارای کد رهگیری و بسته‌بندی ایمن ویژه."
            ),
            100.0: (
                "شمش ۱۰۰ گرمی؛ اوج سرمایه‌گذاری در طلای خالص\n\n"
                "شمش ۱۰۰ گرمی بالاترین سطح سرمایه‌گذاری در طلای فیزیکی را نمایندگی می‌کند. "
                "این وزن با کمترین اجرت ساخت نسبت به گرم، بیشترین بازدهی را برای سرمایه‌گذاران فراهم می‌سازد.\n\n"
                "حداقل اجرت: پایین‌ترین هزینه ساخت به ازای هر گرم طلا.\n"
                "حفظ ثروت: ذخیره ایمن سرمایه‌های بسیار بزرگ در قالبی فشرده.\n"
                "اعتبار و اطمینان: انتخاب نهادهای مالی و سرمایه‌گذاران نهادی.\n\n"
                "با خلوص ۹۹۵ (۲۴ عیار)، هر شمش دارای شناسه یکتا، گواهی اصالت و بسته‌بندی ایمن ویژه است."
            ),
        }

        silver_descriptions = {
            0.1: (
                "شمش نقره ۱۰۰ سوتی؛ آغاز سرمایه‌گذاری در فلز گران‌بها\n\n"
                "شمش نقره ۱۰۰ سوتی کوچک‌ترین واحد سرمایه‌گذاری در نقره خالص است. "
                "نقره به عنوان دومین فلز گران‌بهای جهان، فرصتی منحصربه‌فرد برای تنوع‌بخشی به سبد سرمایه فراهم می‌کند.\n\n"
                "شروع آسان: ورود به بازار نقره با کمترین سرمایه اولیه.\n"
                "هدیه‌ای خاص: گزینه‌ای متفاوت و ارزشمند برای مناسبت‌ها.\n"
                "پتانسیل رشد: نقره با کاربردهای صنعتی گسترده، پتانسیل رشد بالایی دارد.\n\n"
                "هر قطعه با خلوص ۹۹۹.۹ و کد رهگیری اختصاصی عرضه می‌شود."
            ),
            0.2: (
                "شمش نقره ۲۰۰ سوتی؛ پس‌انداز هوشمند در نقره\n\n"
                "شمش نقره ۲۰۰ سوتی فرصتی عالی برای پس‌انداز تدریجی در فلزات گران‌بهاست. "
                "نقره خالص با کاربردهای گسترده صنعتی و سرمایه‌ای، همواره جایگاه ویژه‌ای در سبد دارایی دارد.\n\n"
                "سرمایه‌گذاری تدریجی: خرید منظم و ساخت ذخیره نقره.\n"
                "تنوع سبد: مکمل مناسب برای سرمایه‌گذاری طلا.\n"
                "هدیه‌ای ماندگار: زیبا و ارزشمند برای هر مناسبتی.\n\n"
                "با خلوص ۹۹۹.۹ و بسته‌بندی امن، اصالت هر شمش تضمین شده است."
            ),
            0.5: (
                "شمش نقره نیم گرمی؛ ارزش واقعی در وزنی سبک\n\n"
                "شمش نقره ۵۰۰ سوتی (نیم گرم) یکی از پرطرفدارترین وزن‌ها برای سرمایه‌گذاری خُرد در نقره است. "
                "نقره خالص به عنوان فلزی با ارزش ذاتی و کاربرد صنعتی، آینده‌ای روشن دارد.\n\n"
                "سرمایه‌گذاری هدفمند: قدمی معنادار در مسیر ذخیره نقره.\n"
                "هدیه برازنده: مناسب عروسی، نامزدی و جشن‌های خانوادگی.\n"
                "نقدشوندگی: فروش آسان با قیمت روز بازار.\n\n"
                "هر قطعه با خلوص ۹۹۹.۹، کد رهگیری و بسته‌بندی ایمن ارائه می‌شود."
            ),
            1.0: (
                "شمش نقره یک گرمی؛ استاندارد سرمایه‌گذاری خُرد\n\n"
                "شمش نقره یک گرمی واحد استاندارد سرمایه‌گذاری در نقره خالص است. "
                "نقره با نقش دوگانه خود به‌عنوان فلز گران‌بها و فلز صنعتی، یکی از هوشمندانه‌ترین "
                "گزینه‌های سرمایه‌گذاری محسوب می‌شود.\n\n"
                "ارزش پایدار: نقره همواره به عنوان ذخیره ارزش شناخته شده است.\n"
                "کاربرد صنعتی: تقاضای رو به رشد در صنایع الکترونیک و انرژی خورشیدی.\n"
                "هدیه‌ای بی‌نظیر: زیبا، ماندگار و دارای ارزش واقعی.\n\n"
                "با خلوص ۹۹۹.۹ و گواهی اصالت، هر شمش تحت نظارت دقیق تولید می‌شود."
            ),
            2.5: (
                "شمش نقره ۲.۵ گرمی؛ انتخاب هوشمندانه تنوع‌بخشی\n\n"
                "شمش نقره ۲.۵ گرمی فرصتی عالی برای تنوع‌بخشی به سبد فلزات گران‌بهاست. "
                "ترکیب نقره با طلا در سبد سرمایه، ریسک را کاهش و پتانسیل سود را افزایش می‌دهد.\n\n"
                "تنوع سبد: مکمل ایده‌آل سرمایه‌گذاری طلا.\n"
                "پتانسیل رشد: نقره نسبت به طلا پتانسیل رشد درصدی بالاتری دارد.\n"
                "هدیه‌ای لوکس: انتخابی متفاوت و شایسته برای مناسبت‌های ویژه.\n\n"
                "هر شمش با خلوص ۹۹۹.۹، شناسه یکتا و بسته‌بندی امن عرضه می‌شود."
            ),
            5.0: (
                "شمش نقره ۵ گرمی؛ ذخیره ارزش پایدار\n\n"
                "شمش نقره ۵ گرمی انتخاب محبوب سرمایه‌گذارانی است که به ارزش بلندمدت نقره ایمان دارند. "
                "رشد تقاضای صنعتی نقره در فناوری‌های نوین، آینده‌ای درخشان برای این فلز ترسیم کرده است.\n\n"
                "سرمایه‌گذاری جدی: حجم مناسب برای شروع ذخیره‌سازی نقره.\n"
                "رشد صنعتی: افزایش تقاضا در صنایع فناوری و انرژی سبز.\n"
                "نقدشوندگی: فروش آسان و سریع در بازار.\n\n"
                "تولید شده با خلوص ۹۹۹.۹ و دارای کد رهگیری و گواهی اصالت."
            ),
            10.0: (
                "شمش نقره ۱۰ گرمی؛ سرمایه‌گذاری حرفه‌ای در نقره\n\n"
                "شمش نقره ۱۰ گرمی وزنی استاندارد در بازار نقره است و برای سرمایه‌گذاران جدی طراحی شده. "
                "نقره خالص در کنار طلا، ستون دوم سبد فلزات گران‌بها محسوب می‌شود.\n\n"
                "وزن استاندارد: شناخته‌شده و قابل معامله در بازارهای معتبر.\n"
                "حفظ ارزش: محافظت از سرمایه در برابر تورم.\n"
                "بازدهی بالقوه: پتانسیل رشد قیمت نقره در بلندمدت.\n\n"
                "با خلوص ۹۹۹.۹، هر شمش دارای شناسه یکتا و گواهی اصالت است."
            ),
            20.0: (
                "شمش نقره ۲۰ گرمی؛ گنجینه نقره‌ای\n\n"
                "شمش نقره ۲۰ گرمی ذخیره‌ای ارزشمند برای سرمایه‌گذاران بلندمدت است. "
                "با افزایش کاربردهای صنعتی نقره، این فلز همچنان یکی از جذاب‌ترین گزینه‌های سرمایه‌گذاری باقی مانده است.\n\n"
                "سرمایه‌گذاری بلندمدت: حفظ ثروت در قالب فلز گران‌بها.\n"
                "تقاضای رو به رشد: افزایش مصرف صنعتی نقره در جهان.\n"
                "نقدشوندگی تضمینی: فروش آسان با قیمت بازار.\n\n"
                "هر قطعه با خلوص ۹۹۹.۹ و بسته‌بندی ایمن و کد رهگیری عرضه می‌شود."
            ),
            31.1: (
                "شمش نقره یک اونسی؛ استاندارد بین‌المللی نقره\n\n"
                "شمش نقره یک اونس (۳۱.۱ گرم) واحد اصلی قیمت‌گذاری نقره در بازارهای جهانی است. "
                "خرید نقره اونسی، سرمایه‌گذاری مطابق با استانداردهای بین‌المللی است.\n\n"
                "مرجع قیمت: قیمت نقره در بازارهای جهانی بر اساس اونس تعیین می‌شود.\n"
                "سرمایه‌گذاری حرفه‌ای: انتخاب سرمایه‌گذاران باتجربه و نهادهای مالی.\n"
                "نقدشوندگی بالا: بالاترین سطح تقاضا و سهولت معامله.\n\n"
                "با خلوص ۹۹۹.۹ و گواهی اصالت، هر شمش اونسی با شناسه یکتا عرضه می‌شود."
            ),
            50.0: (
                "شمش نقره ۵۰ گرمی؛ ذخیره سنگین فلز گران‌بها\n\n"
                "شمش نقره ۵۰ گرمی برای سرمایه‌گذارانی طراحی شده که حجم بالایی از نقره خالص را ذخیره می‌کنند. "
                "این وزن با اجرت ساخت بهینه‌تر، بازدهی بهتری نسبت به وزن‌های سبک‌تر ارائه می‌دهد.\n\n"
                "بهینه‌سازی هزینه: اجرت ساخت کمتر به ازای هر گرم.\n"
                "سرمایه‌گذاری کلان: ذخیره حجم بالای نقره خالص.\n"
                "ارزش صنعتی: نقره با تقاضای رو به رشد صنعتی، آینده‌ای روشن دارد.\n\n"
                "تولید شده با خلوص ۹۹۹.۹، دارای کد رهگیری و بسته‌بندی ایمن ویژه."
            ),
            100.0: (
                "شمش نقره ۱۰۰ گرمی؛ اوج سرمایه‌گذاری نقره‌ای\n\n"
                "شمش نقره ۱۰۰ گرمی بالاترین سطح سرمایه‌گذاری در نقره فیزیکی را نمایندگی می‌کند. "
                "با کمترین اجرت ساخت و بالاترین کارایی، این وزن انتخاب سرمایه‌گذاران نهادی است.\n\n"
                "حداقل اجرت: پایین‌ترین هزینه ساخت به ازای هر گرم نقره.\n"
                "حفظ ثروت: ذخیره امن سرمایه‌های بزرگ.\n"
                "تقاضای جهانی: نقره همواره جزو پرتقاضاترین فلزات گران‌بها بوده است.\n\n"
                "با خلوص ۹۹۹.۹، هر شمش دارای شناسه یکتا، گواهی اصالت و بسته‌بندی ایمن ویژه است."
            ),
        }

        # Product type definitions (folder → Excel sheet mapping)
        product_types = [
            {
                "folder": "شمش طلا با بسته بندی",
                "excel_sheet": "شمش طلاملا",
                "slug": "gold-talamala",
                "category_slug": "gold-talamala",
                "name_prefix": "شمش طلا طلاملا",
                "purity": 995,
                "metal": "gold",
            },
            {
                "folder": "شمش طلا بدون بسته بندی",
                "excel_sheet": "شمش سرمایه ای",
                "slug": "gold-investment",
                "category_slug": "gold-investment",
                "name_prefix": "شمش طلا سرمایه‌ای",
                "purity": 995,
                "metal": "gold",
            },
            {
                "folder": "شمش نقره با بسته بندی",
                "excel_sheet": "شمش نقره طلاملا",
                "slug": "silver-talamala",
                "category_slug": "silver-talamala",
                "name_prefix": "شمش نقره طلاملا",
                "purity": 999.9,
                "metal": "silver",
            },
            {
                "folder": "شمش نقره بدون بسته بندی",
                "excel_sheet": "شمش سرمایه ای نقره",
                "slug": "silver-investment",
                "category_slug": "silver-investment",
                "name_prefix": "شمش نقره سرمایه‌ای",
                "purity": 999.9,
                "metal": "silver",
            },
        ]

        # Weight definitions: (grams, persian_label, {folder overrides for non-standard names})
        weights_def = [
            (0.1,  "۱۰۰ سوت",  {"شمش نقره با بسته بندی": "0.1 g"}),
            (0.2,  "۲۰۰ سوت",  {}),
            (0.5,  "۵۰۰ سوت",  {}),
            (1.0,  "۱ گرم",     {}),
            (2.5,  "۲.۵ گرم",   {}),
            (5.0,  "۵ گرم",     {}),
            (10.0, "۱۰ گرم",    {}),
            (20.0, "۲۰ گرم",    {}),
            (31.1, "۱ اونس",    {
                "شمش طلا با بسته بندی": "1ounce",
                "شمش طلا بدون بسته بندی": "1ounce",
                "شمش نقره با بسته بندی": "1onuce",
                "شمش نقره بدون بسته بندی": "1onuce",
            }),
            (50.0, "۵۰ گرم",    {"شمش طلا بدون بسته بندی": "50G"}),
            (100.0,"۱۰۰ گرم",   {}),
        ]

        def weight_to_folder(weight_grams, ptype_folder, overrides):
            """Convert weight to image folder name."""
            if ptype_folder in overrides:
                return overrides[ptype_folder]
            if weight_grams == 31.1:
                return "1ounce"
            if weight_grams == int(weight_grams):
                return f"{int(weight_grams)}g"
            return f"{weight_grams}g"

        # --- Read wage data from Excel file ---
        # Excel format: col A = weight (grams), B-E = tier wages (decimal → × 100 = percent)
        EXCEL_PATH = os.path.join(IMG_SRC_BASE, "سطوح قیمت گذاری طلا ونقره.xlsx")
        wage_data = {}  # {type_slug: {weight: [tier1%, tier2%, tier3%, end_customer%]}}

        if os.path.exists(EXCEL_PATH):
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
            for ptype in product_types:
                sheet_name = ptype["excel_sheet"]
                if sheet_name not in wb.sheetnames:
                    print(f"  ⚠ Excel sheet '{sheet_name}' not found for {ptype['slug']}")
                    continue
                ws = wb[sheet_name]
                slug_wages = {}
                for row in ws.iter_rows(min_row=3, max_col=5, values_only=True):
                    weight_val, t1, t2, t3, t4 = row
                    if weight_val is None or t4 is None:
                        continue
                    w = round(float(weight_val), 2)
                    # Excel stores as decimal (e.g. 0.14), convert to percent (14.0)
                    slug_wages[w] = [
                        round(float(t1 or 0) * 100, 2),
                        round(float(t2 or 0) * 100, 2),
                        round(float(t3 or 0) * 100, 2),
                        round(float(t4 or 0) * 100, 2),
                    ]
                wage_data[ptype["slug"]] = slug_wages
            wb.close()
            print(f"  Wage data loaded from Excel ({len(wage_data)} types)")
            for slug, wd in wage_data.items():
                print(f"    {slug}: {len(wd)} weights")
        else:
            print(f"  ⚠ Excel file not found: {EXCEL_PATH}")
            print("    Products will be created without wage data!")

        print(f"\n[5/9] Products ({len(weights_def)} weights x {len(product_types)} types)")

        product_map = {}
        default_design = db.query(CardDesign).first()
        default_package = db.query(PackageType).first()

        for ptype in product_types:
            cat = cat_map.get(ptype["category_slug"])
            type_wages = wage_data.get(ptype["slug"], {})
            desc_map = gold_descriptions if ptype["metal"] == "gold" else silver_descriptions

            for weight_grams, weight_label, folder_overrides in weights_def:
                name = f"{ptype['name_prefix']} {weight_label}"
                tier_wages_list = type_wages.get(weight_grams, [0, 0, 0, 0])
                end_customer_wage = tier_wages_list[3]
                description = desc_map.get(weight_grams)

                existing = db.query(Product).filter(Product.name == name).first()
                if existing:
                    p = existing
                    product_map[name] = p

                    # Update wage from Excel
                    if end_customer_wage and p.wage != end_customer_wage:
                        p.wage = end_customer_wage
                        print(f"  ~ wage updated: {name} → {end_customer_wage}%")

                    # Update metal_type
                    expected_metal = ptype.get("metal", "gold")
                    if p.metal_type != expected_metal:
                        p.metal_type = expected_metal

                    # Update description
                    if description:
                        p.description = description

                    # Refresh images: delete old ones, copy new ones
                    old_images = db.query(ProductImage).filter(ProductImage.product_id == p.id).all()
                    for old_img in old_images:
                        old_file = os.path.join(PROJECT_ROOT, old_img.file_path)
                        if os.path.exists(old_file):
                            os.remove(old_file)
                        db.delete(old_img)
                    db.flush()
                    print(f"  ~ refreshing images: {name}")
                else:
                    p = Product(
                        name=name,
                        weight=weight_grams,
                        purity=ptype["purity"],
                        metal_type=ptype.get("metal", "gold"),
                        card_design_id=default_design.id if default_design else None,
                        package_type_id=default_package.id if default_package else None,
                        wage=end_customer_wage,
                        is_wage_percent=True,
                        is_active=True,
                        description=description,
                    )
                    db.add(p)
                    db.flush()

                    # M2M category link
                    if cat:
                        db.add(ProductCategoryLink(product_id=p.id, category_id=cat.id))

                    print(f"  + {name} ({weight_grams}g, purity={ptype['purity']}, wage={end_customer_wage}%)")

                # Copy images from _private to static/uploads/products/ (max 2: front + back)
                folder_name = weight_to_folder(weight_grams, ptype["folder"], folder_overrides)
                img_src_dir = os.path.join(IMG_SRC_BASE, ptype["folder"], folder_name)
                if os.path.isdir(img_src_dir):
                    src_files = []
                    for root, _dirs, files in os.walk(img_src_dir):
                        for f in sorted(files):
                            if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                                src_files.append(os.path.join(root, f))
                    src_files.sort()
                    # Only keep first 2 images (front + back)
                    src_files = src_files[:2]

                    for idx, src_path in enumerate(src_files):
                        ext = os.path.splitext(src_path)[1].lower()
                        clean_name = f"{ptype['slug']}_{folder_name}_{idx + 1}{ext}"
                        dst_path = os.path.join(IMG_DST_BASE, clean_name)
                        shutil.copy2(src_path, dst_path)
                        rel_path = f"static/uploads/products/{clean_name}"
                        db.add(ProductImage(
                            product_id=p.id,
                            file_path=rel_path,
                            is_default=(idx == 0),
                        ))

                product_map[name] = p

        db.flush()

        # ==========================================
        # 6. (Removed — locations merged into dealers)
        # ==========================================

        # ==========================================
        # 7. Bars (inventory) — created AFTER dealers (section 10)
        # ==========================================
        print("\n[6/9] Bars (deferred — created after dealers)")

        # (Bars will be created in section 10.5 after dealers exist)
        _bar_batch1 = db.query(Batch).filter(Batch.batch_number == "B-1403-001").first()

        # ==========================================
        # 8. Sample Coupons
        # ==========================================
        print("\n[8/9] Coupons")

        coupons_data = [
            {
                "code": "WELCOME10", "title": "خوش‌آمدگویی ۱۰٪ تخفیف",
                "description": "تخفیف ویژه اولین خرید",
                "coupon_type": "DISCOUNT", "discount_mode": "PERCENT",
                "discount_value": 10, "max_discount_amount": 50_000_000,
                "first_purchase_only": True, "max_per_customer": 1, "max_total_uses": 100,
            },
            {
                "code": "CASHBACK5", "title": "۵٪ کشبک",
                "description": "۵ درصد مبلغ خرید به کیف پول واریز می‌شود",
                "coupon_type": "CASHBACK", "discount_mode": "PERCENT",
                "discount_value": 5, "max_discount_amount": 20_000_000,
                "max_per_customer": 3,
            },
            {
                "code": "FIXED500", "title": "۵۰۰ هزار تومان تخفیف",
                "description": "تخفیف ثابت برای خرید بالای ۵ میلیون",
                "coupon_type": "DISCOUNT", "discount_mode": "FIXED",
                "discount_value": 5_000_000, "min_order_amount": 50_000_000,
                "max_per_customer": 1, "max_total_uses": 50,
            },
            {
                "code": "VIP2026", "title": "کد ویژه VIP (موبایل خاص)",
                "description": "تخفیف ۱۵٪ برای مشتریان ویژه",
                "coupon_type": "DISCOUNT", "discount_mode": "PERCENT",
                "discount_value": 15, "max_discount_amount": 100_000_000,
                "is_private": True, "max_per_customer": 5,
                "_mobiles": ["09351234567", "09359876543"],
            },
            {
                "code": "GOLD10", "title": "۱۰٪ تخفیف شمش طلا",
                "description": "فقط برای دسته شمش طلا",
                "coupon_type": "DISCOUNT", "discount_mode": "PERCENT",
                "discount_value": 10, "max_discount_amount": 80_000_000,
                "scope": "CATEGORY", "max_per_customer": 2,
                "_category_slugs": ["gold-talamala", "gold-investment"],
            },
        ]

        for cd in coupons_data:
            if db.query(Coupon).filter(Coupon.code == cd["code"]).first():
                print(f"  = exists: {cd['code']}")
                continue

            mobiles = cd.pop("_mobiles", [])
            cat_slugs = cd.pop("_category_slugs", [])
            coupon = Coupon(
                code=cd["code"], title=cd["title"],
                description=cd.get("description", ""),
                coupon_type=cd.get("coupon_type", "DISCOUNT"),
                discount_mode=cd.get("discount_mode", "PERCENT"),
                discount_value=cd["discount_value"],
                max_discount_amount=cd.get("max_discount_amount"),
                scope=cd.get("scope", "GLOBAL"),
                min_order_amount=cd.get("min_order_amount", 0),
                first_purchase_only=cd.get("first_purchase_only", False),
                is_private=cd.get("is_private", False),
                max_per_customer=cd.get("max_per_customer", 1),
                max_total_uses=cd.get("max_total_uses"),
                status="ACTIVE",
            )
            db.add(coupon)
            db.flush()
            for m in mobiles:
                db.add(CouponMobile(coupon_id=coupon.id, mobile=m, note="Seed VIP"))
            for slug in cat_slugs:
                cat = db.query(ProductCategory).filter(ProductCategory.slug == slug).first()
                if cat:
                    db.add(CouponCategory(coupon_id=coupon.id, category_id=cat.id))
            db.flush()
            tag = f" [mobile: {', '.join(mobiles)}]" if mobiles else ""
            if cat_slugs:
                tag += f" [categories: {', '.join(cat_slugs)}]"
            print(f"  + {cd['code']}: {cd['title']}{tag}")

        # ==========================================
        # 8.5 Geographic Data (Provinces / Cities / Districts)
        # ==========================================
        print("\n[8.5] Geographic Data")

        geo_data = {
            "تهران": {
                "cities": ["تهران", "شهریار", "اسلامشهر", "ورامین", "ری", "قدس", "پاکدشت", "ملارد", "رباط‌کریم", "بومهن", "دماوند"],
                "districts": {
                    "تهران": ["همه محله‌ها", "تهرانپارس", "نارمک", "پیروزی", "ونک", "سعادت‌آباد", "شهرک غرب",
                              "پونک", "ستارخان", "آریاشهر", "جنت‌آباد", "اکباتان", "تجریش", "ولیعصر",
                              "میرداماد", "یوسف‌آباد", "امیرآباد", "انقلاب", "بازار", "مولوی", "جمهوری",
                              "فردوسی", "آزادی", "صادقیه", "المهدی", "چیتگر", "اندرزگو", "زعفرانیه"],
                },
            },
            "اصفهان": {
                "cities": ["اصفهان", "کاشان", "نجف‌آباد", "خمینی‌شهر", "شاهین‌شهر", "فلاورجان", "لنجان"],
                "districts": {"اصفهان": ["همه محله‌ها", "چهارباغ", "جلفا", "خوراسگان", "شاهین‌ویلا", "ملک‌شهر"]},
            },
            "فارس": {
                "cities": ["شیراز", "مرودشت", "کازرون", "فسا", "جهرم", "لار", "داراب", "آباده"],
                "districts": {"شیراز": ["همه محله‌ها", "زند", "معالی‌آباد", "قصرالدشت", "ملاصدرا", "صدرا"]},
            },
            "خراسان رضوی": {
                "cities": ["مشهد", "نیشابور", "سبزوار", "تربت‌حیدریه", "قوچان", "کاشمر", "گناباد", "چناران"],
                "districts": {"مشهد": ["همه محله‌ها", "وکیل‌آباد", "احمدآباد", "هاشمیه", "الهیه", "سجاد"]},
            },
            "آذربایجان شرقی": {
                "cities": ["تبریز", "مراغه", "مرند", "میانه", "اهر", "سراب", "بناب", "شبستر"],
                "districts": {"تبریز": ["همه محله‌ها", "ولیعصر", "آزادی", "ائل‌گلی", "باغمیشه", "رشدیه"]},
            },
            "خوزستان": {
                "cities": ["اهواز", "آبادان", "خرمشهر", "دزفول", "شوشتر", "ماهشهر", "بهبهان", "اندیمشک"],
                "districts": {"اهواز": ["همه محله‌ها", "کیانپارس", "گلستان", "زیتون", "پادادشهر"]},
            },
            "گیلان": {
                "cities": ["رشت", "بندر انزلی", "لاهیجان", "لنگرود", "آستارا", "تالش", "فومن", "صومعه‌سرا"],
                "districts": {"رشت": ["همه محله‌ها"]},
            },
            "مازندران": {
                "cities": ["ساری", "بابل", "آمل", "قائم‌شهر", "بهشهر", "تنکابن", "رامسر", "نوشهر", "چالوس", "نور"],
                "districts": {"ساری": ["همه محله‌ها"]},
            },
            "کرمان": {
                "cities": ["کرمان", "رفسنجان", "جیرفت", "سیرجان", "بم", "زرند", "بردسیر"],
                "districts": {"کرمان": ["همه محله‌ها"]},
            },
            "البرز": {
                "cities": ["کرج", "فردیس", "نظرآباد", "هشتگرد", "محمدشهر", "مشکین‌دشت"],
                "districts": {"کرج": ["همه محله‌ها", "گوهردشت", "مهرشهر", "عظیمیه", "گلشهر", "جهانشهر"]},
            },
            "قم": {
                "cities": ["قم"],
                "districts": {"قم": ["همه محله‌ها"]},
            },
            "کرمانشاه": {
                "cities": ["کرمانشاه", "اسلام‌آباد غرب", "سنقر", "کنگاور", "پاوه", "جوانرود", "هرسین"],
                "districts": {"کرمانشاه": ["همه محله‌ها"]},
            },
            "آذربایجان غربی": {
                "cities": ["ارومیه", "خوی", "مهاباد", "بوکان", "میاندوآب", "سلماس", "پیرانشهر", "نقده"],
                "districts": {"ارومیه": ["همه محله‌ها"]},
            },
            "هرمزگان": {
                "cities": ["بندرعباس", "قشم", "کیش", "میناب", "بندر لنگه", "حاجی‌آباد"],
                "districts": {"بندرعباس": ["همه محله‌ها"]},
            },
            "سیستان و بلوچستان": {
                "cities": ["زاهدان", "چابهار", "ایرانشهر", "سراوان", "زابل", "خاش", "نیک‌شهر"],
                "districts": {"زاهدان": ["همه محله‌ها"]},
            },
            "لرستان": {
                "cities": ["خرم‌آباد", "بروجرد", "دورود", "الیگودرز", "کوهدشت", "نورآباد"],
                "districts": {"خرم‌آباد": ["همه محله‌ها"]},
            },
            "همدان": {
                "cities": ["همدان", "ملایر", "نهاوند", "تویسرکان", "بهار", "اسدآباد", "کبودرآهنگ"],
                "districts": {"همدان": ["همه محله‌ها"]},
            },
            "مرکزی": {
                "cities": ["اراک", "ساوه", "خمین", "دلیجان", "محلات", "شازند", "تفرش"],
                "districts": {"اراک": ["همه محله‌ها"]},
            },
            "گلستان": {
                "cities": ["گرگان", "گنبد کاووس", "علی‌آباد کتول", "بندر ترکمن", "آق‌قلا", "مینودشت"],
                "districts": {"گرگان": ["همه محله‌ها"]},
            },
            "اردبیل": {
                "cities": ["اردبیل", "پارس‌آباد", "خلخال", "مشگین‌شهر", "نمین", "نیر"],
                "districts": {"اردبیل": ["همه محله‌ها"]},
            },
            "یزد": {
                "cities": ["یزد", "میبد", "اردکان", "تفت", "بافق", "ابرکوه", "مهریز"],
                "districts": {"یزد": ["همه محله‌ها"]},
            },
            "زنجان": {
                "cities": ["زنجان", "ابهر", "خدابنده", "خرمدره", "ماهنشان", "طارم"],
                "districts": {"زنجان": ["همه محله‌ها"]},
            },
            "سمنان": {
                "cities": ["سمنان", "شاهرود", "دامغان", "گرمسار", "مهدی‌شهر"],
                "districts": {"سمنان": ["همه محله‌ها"]},
            },
            "قزوین": {
                "cities": ["قزوین", "تاکستان", "بویین‌زهرا", "آبیک", "آوج", "الموت"],
                "districts": {"قزوین": ["همه محله‌ها"]},
            },
            "بوشهر": {
                "cities": ["بوشهر", "برازجان", "کنگان", "گناوه", "دیلم", "دیر", "جم"],
                "districts": {"بوشهر": ["همه محله‌ها"]},
            },
            "کهگیلویه و بویراحمد": {
                "cities": ["یاسوج", "دهدشت", "گچساران", "دوگنبدان", "لیکک"],
                "districts": {"یاسوج": ["همه محله‌ها"]},
            },
            "چهارمحال و بختیاری": {
                "cities": ["شهرکرد", "بروجن", "فارسان", "لردگان", "اردل", "کیار"],
                "districts": {"شهرکرد": ["همه محله‌ها"]},
            },
            "کردستان": {
                "cities": ["سنندج", "سقز", "مریوان", "بانه", "قروه", "بیجار", "دیواندره", "کامیاران"],
                "districts": {"سنندج": ["همه محله‌ها"]},
            },
            "ایلام": {
                "cities": ["ایلام", "دهلران", "ایوان", "آبدانان", "مهران", "دره‌شهر"],
                "districts": {"ایلام": ["همه محله‌ها"]},
            },
            "خراسان شمالی": {
                "cities": ["بجنورد", "شیروان", "اسفراین", "جاجرم", "فاروج"],
                "districts": {"بجنورد": ["همه محله‌ها"]},
            },
            "خراسان جنوبی": {
                "cities": ["بیرجند", "قاین", "طبس", "فردوس", "نهبندان", "سربیشه"],
                "districts": {"بیرجند": ["همه محله‌ها"]},
            },
        }

        existing_prov = db.query(GeoProvince).count()
        if existing_prov == 0:
            sort_idx = 0
            for prov_name, prov_info in geo_data.items():
                sort_idx += 1
                prov = GeoProvince(name=prov_name, sort_order=sort_idx)
                db.add(prov)
                db.flush()

                for city_name in prov_info["cities"]:
                    city = GeoCity(province_id=prov.id, name=city_name)
                    db.add(city)
                    db.flush()

                    district_list = prov_info.get("districts", {}).get(city_name, [])
                    for dist_name in district_list:
                        db.add(GeoDistrict(city_id=city.id, name=dist_name))

            db.flush()
            total_cities = db.query(GeoCity).count()
            total_districts = db.query(GeoDistrict).count()
            print(f"  + {len(geo_data)} provinces, {total_cities} cities, {total_districts} districts")
        else:
            print(f"  = geo data exists ({existing_prov} provinces)")

        # ==========================================
        # 9. Wallet: Credit test customer
        # ==========================================
        print("\n[9/9] Test Wallet Credit")

        test_customer = db.query(User).filter(User.mobile == "09351234567").first()
        if test_customer:
            existing_acct = db.query(Account).filter(
                Account.user_id == test_customer.id,
                Account.asset_code == "IRR",
            ).first()
            if not existing_acct:
                acct = Account(
                    user_id=test_customer.id,
                    asset_code="IRR",
                    balance=100_000_000, locked_balance=0,
                )
                db.add(acct)
                db.flush()
                db.add(LedgerEntry(
                    account_id=acct.id, txn_type="Deposit",
                    delta_balance=100_000_000, delta_locked=0,
                    balance_after=100_000_000, locked_after=0,
                    idempotency_key="seed:initial_credit:1",
                    reference_type="seed", reference_id="initial",
                    description="شارژ اولیه تستی",
                ))
                print(f"  + {test_customer.full_name} -> 10,000,000 toman (IRR)")
            else:
                print(f"  = wallet exists for {test_customer.full_name}")

            # Gold wallet for test customer
            existing_gold = db.query(Account).filter(
                Account.user_id == test_customer.id,
                Account.asset_code == "XAU_MG",
            ).first()
            if not existing_gold:
                db.add(Account(
                    user_id=test_customer.id,
                    asset_code="XAU_MG",
                    balance=0, locked_balance=0,
                ))
                print(f"  + {test_customer.full_name} -> Gold wallet (XAU_MG)")
            else:
                print(f"  = gold wallet exists for {test_customer.full_name}")

        # ==========================================
        # 9.1 Sample Withdrawal Requests
        # ==========================================
        print("\n[9.1] Sample Withdrawal Requests")

        existing_wr = db.query(WithdrawalRequest).count()
        if existing_wr == 0 and test_customer:
            from modules.wallet.models import WithdrawalStatus

            acct = db.query(Account).filter(
                Account.user_id == test_customer.id,
                Account.asset_code == "IRR",
            ).first()

            # --- Customer withdrawals ---
            wr_data = [
                {
                    "user_id": test_customer.id,
                    "amount_irr": 5_000_000,
                    "shaba_number": "IR820540102680020817909002",
                    "account_holder": "علی رضایی",
                    "status": WithdrawalStatus.PENDING,
                },
                {
                    "user_id": test_customer.id,
                    "amount_irr": 10_000_000,
                    "shaba_number": "IR062960000000100324200001",
                    "account_holder": "علی رضایی",
                    "status": WithdrawalStatus.PENDING,
                },
                {
                    "user_id": test_customer.id,
                    "amount_irr": 3_000_000,
                    "shaba_number": "IR820540102680020817909002",
                    "account_holder": "علی رضایی",
                    "status": WithdrawalStatus.PAID,
                    "admin_note": "واریز شد - شماره پیگیری ۱۲۳۴۵۶",
                },
                {
                    "user_id": test_customer.id,
                    "amount_irr": 2_000_000,
                    "shaba_number": "IR062960000000100324200001",
                    "account_holder": "علی رضایی",
                    "status": WithdrawalStatus.REJECTED,
                    "admin_note": "شماره شبا با نام صاحب حساب مطابقت ندارد",
                },
            ]

            # --- Dealer withdrawal (test for dealer withdrawal feature) ---
            test_dealer_wr = db.query(User).filter(User.mobile == "09161234567", User.is_dealer == True).first()
            if test_dealer_wr:
                wr_data.append({
                    "user_id": test_dealer_wr.id,
                    "amount_irr": 8_000_000,
                    "shaba_number": "IR550170000000100000000007",
                    "account_holder": "احمد نوری",
                    "status": WithdrawalStatus.PENDING,
                    "_is_dealer": True,
                })

            total_pending_hold = {}  # {user_id: amount}
            for wd in wr_data:
                wr = WithdrawalRequest(
                    user_id=wd["user_id"],
                    amount_irr=wd["amount_irr"],
                    shaba_number=wd["shaba_number"],
                    account_holder=wd["account_holder"],
                    status=wd["status"],
                    admin_note=wd.get("admin_note"),
                )
                db.add(wr)
                label = "dealer" if wd.get("_is_dealer") else "customer"
                status_label = wd["status"].value if hasattr(wd["status"], "value") else wd["status"]
                print(f"  + Withdrawal {wd['amount_irr'] // 10:,} toman [{label}] [{status_label}]")

                if wd["status"] == WithdrawalStatus.PENDING:
                    total_pending_hold[wd["user_id"]] = (
                        total_pending_hold.get(wd["user_id"], 0) + wd["amount_irr"]
                    )

            # Hold funds for pending withdrawals
            for uid, hold_amount in total_pending_hold.items():
                user_acct = db.query(Account).filter(
                    Account.user_id == uid,
                    Account.asset_code == "IRR",
                ).first()
                if user_acct:
                    user_acct.locked_balance += hold_amount
                    db.flush()
                    db.add(LedgerEntry(
                        account_id=user_acct.id, txn_type="Hold",
                        delta_balance=0, delta_locked=hold_amount,
                        balance_after=user_acct.balance, locked_after=user_acct.locked_balance,
                        idempotency_key=f"seed:withdrawal_hold:user:{uid}",
                        reference_type="seed", reference_id="withdrawal_hold",
                        description="بلوکه تستی برای درخواست‌های برداشت",
                    ))
                    print(f"  + Held {hold_amount // 10:,} toman for user #{uid} pending withdrawals")

            db.flush()
        else:
            print(f"  = {existing_wr} withdrawal requests exist, skipping")

        # ==========================================
        # 9.5. Dealer Tiers
        # ==========================================
        print("\n[9.5] Dealer Tiers")

        tiers_data = [
            {"name": "پخش", "slug": "distributor", "sort_order": 1, "is_end_customer": False},
            {"name": "بنکدار", "slug": "wholesaler", "sort_order": 2, "is_end_customer": False},
            {"name": "فروشگاه", "slug": "store", "sort_order": 3, "is_end_customer": False},
            {"name": "مشتری نهایی", "slug": "end_customer", "sort_order": 4, "is_end_customer": True},
        ]

        tier_map = {}
        for td in tiers_data:
            existing = db.query(DealerTier).filter(DealerTier.slug == td["slug"]).first()
            if not existing:
                tier = DealerTier(**td, is_active=True)
                db.add(tier)
                db.flush()
                tier_map[td["slug"]] = tier
                ec = " [END-CUSTOMER]" if td["is_end_customer"] else ""
                print(f"  + {td['name']} ({td['slug']}){ec}")
            else:
                tier_map[td["slug"]] = existing
                print(f"  = exists: {td['name']}")

        db.flush()

        # ==========================================
        # 9.6. Product Tier Wages
        # ==========================================
        print("\n[9.6] Product Tier Wages")

        # Wage data from Excel (same as section 5, reused for tier wages)
        # {type_slug: {weight: [distributor%, wholesaler%, store%, end_customer%]}}
        tier_slugs_order = ["distributor", "wholesaler", "store", "end_customer"]

        # Always refresh tier wages from Excel data
        # Delete existing tier wages and recreate from current Excel
        existing_tw = db.query(ProductTierWage).count()
        if existing_tw > 0:
            db.query(ProductTierWage).delete()
            db.flush()
            print(f"  ~ deleted {existing_tw} old tier wages, recreating from Excel...")

        tw_count = 0
        # Build category_id → type_slug lookup
        cat_id_to_slug = {}
        for slug, cat_obj in cat_map.items():
            cat_id_to_slug[cat_obj.id] = slug

        all_products = db.query(Product).filter(Product.is_active == True).all()
        for p in all_products:
            # Determine product type from category
            type_slug = None
            for cid in p.category_ids:
                if cid in cat_id_to_slug:
                    type_slug = cat_id_to_slug[cid]
                    break
            if not type_slug or type_slug not in wage_data:
                continue

            weight_key = float(p.weight)
            wages = wage_data[type_slug].get(weight_key)
            if not wages:
                continue

            for idx, tier_slug in enumerate(tier_slugs_order):
                tier = tier_map.get(tier_slug)
                if tier:
                    db.add(ProductTierWage(
                        product_id=p.id,
                        tier_id=tier.id,
                        wage_percent=wages[idx],
                    ))
                    tw_count += 1
        db.flush()
        print(f"  + {tw_count} tier wages created from Excel")

        # ==========================================
        # 10. Dealers
        # ==========================================
        print("\n[10] Dealers")

        tier_distributor = tier_map.get("distributor")
        tier_wholesaler = tier_map.get("wholesaler")
        tier_store = tier_map.get("store")

        # Lookup geo IDs for dealer addresses
        def geo_ids(province_name, city_name):
            prov = db.query(GeoProvince).filter(GeoProvince.name == province_name).first()
            city = db.query(GeoCity).filter(GeoCity.province_id == prov.id, GeoCity.name == city_name).first() if prov else None
            return (prov.id if prov else None, city.id if city else None)

        geo_tehran = geo_ids("تهران", "تهران")
        geo_esfahan = geo_ids("اصفهان", "اصفهان")
        geo_shiraz = geo_ids("فارس", "شیراز")
        geo_mashhad = geo_ids("خراسان رضوی", "مشهد")
        geo_tabriz = geo_ids("آذربایجان شرقی", "تبریز")

        dealers_data = [
            # --- انبار مرکزی (postal hub + warehouse) ---
            {
                "mobile": "00000000000",
                "full_name": "انبار مرکزی تهران",
                "national_id": "0000000000",
                "api_key": None,
                "tier_id": None,
                "province_id": geo_tehran[0], "city_id": geo_tehran[1],
                "address": "تهران، خیابان ولیعصر، پلاک ۱۲۳",
                "landline_phone": "02188001234",
                "is_warehouse": True,
                "is_postal_hub": True,
            },
            # --- شهرستان ---
            {
                "mobile": "09161234567",
                "full_name": "احمد نوری",
                "national_id": "1111111111",
                "api_key": "test_esfahan_key_0000000000000000",
                "tier_id": tier_distributor.id if tier_distributor else None,
                "province_id": geo_esfahan[0], "city_id": geo_esfahan[1],
                "address": "اصفهان، خیابان چهارباغ، پلاک ۴۵",
                "landline_phone": "03132001234",
            },
            {
                "mobile": "09171234567",
                "full_name": "سارا کریمی",
                "national_id": "2222222222",
                "api_key": "test_shiraz__key_1111111111111111",
                "tier_id": tier_wholesaler.id if tier_wholesaler else None,
                "province_id": geo_shiraz[0], "city_id": geo_shiraz[1],
                "address": "شیراز، خیابان زند، پلاک ۷۸",
                "landline_phone": "07136001234",
            },
            {
                "mobile": "09181234567",
                "full_name": "حسین موسوی",
                "national_id": "3333333333",
                "api_key": "test_mashhad_key_2222222222222222",
                "tier_id": tier_store.id if tier_store else None,
                "province_id": geo_mashhad[0], "city_id": geo_mashhad[1],
                "address": "مشهد، بلوار وکیل‌آباد، پلاک ۲۲",
                "landline_phone": "05138001234",
            },
            {
                "mobile": "09141234567",
                "full_name": "یوسف قربانی",
                "national_id": "4444444444",
                "api_key": "test_tabriz__key_3333333333333333",
                "tier_id": tier_distributor.id if tier_distributor else None,
                "province_id": geo_tabriz[0], "city_id": geo_tabriz[1],
                "address": "تبریز، خیابان آزادی، پلاک ۳۳",
                "landline_phone": "04135001234",
            },
            # --- تهران ---
            {
                "mobile": "09121234567",
                "full_name": "محمد رضایی",
                "national_id": "5555555555",
                "api_key": "test_mirdmad_key_4444444444444444",
                "tier_id": tier_distributor.id if tier_distributor else None,
                "province_id": geo_tehran[0], "city_id": geo_tehran[1],
                "address": "بلوار میرداماد، برج آرین، طبقه دوم اداری، واحد ۵",
                "landline_phone": "02145241",
            },
            {
                "mobile": "09122345678",
                "full_name": "علی حسینی",
                "national_id": "6666666666",
                "api_key": "test_nasrkhr_key_5555555555555555",
                "tier_id": tier_wholesaler.id if tier_wholesaler else None,
                "province_id": geo_tehran[0], "city_id": geo_tehran[1],
                "address": "بازار تهران، خیابان ناصر خسرو، پاساژ شمس العماره، طبقه منفی ۱، واحد ۳۰۵",
                "landline_phone": "02186091012",
            },
            {
                "mobile": "09123456780",
                "full_name": "فاطمه احمدی",
                "national_id": "7777777777",
                "api_key": "test_ordibht_key_6666666666666666",
                "tier_id": tier_wholesaler.id if tier_wholesaler else None,
                "province_id": geo_tehran[0], "city_id": geo_tehran[1],
                "address": "بازار بزرگ تهران، پاساژ اردیبهشت، طبقه هم‌کف، پلاک ۶۸",
                "landline_phone": "02186091013",
            },
            {
                "mobile": "09124567890",
                "full_name": "رضا محمدی",
                "national_id": "8888888888",
                "api_key": "test_shahrak_key_7777777777777777",
                "tier_id": tier_store.id if tier_store else None,
                "province_id": geo_tehran[0], "city_id": geo_tehran[1],
                "address": "بلوار فرحزادی، مجتمع تجاری لیدوما، تجاری دوم (G2)، واحد ۱۵",
                "landline_phone": "02186091014",
            },
            {
                "mobile": "09125678901",
                "full_name": "مریم کاظمی",
                "national_id": "9999999999",
                "api_key": "test_karimkh_key_8888888888888888",
                "tier_id": tier_store.id if tier_store else None,
                "province_id": geo_tehran[0], "city_id": geo_tehran[1],
                "address": "میدان ولی‌عصر(عج)، خیابان کریمخان، مجتمع تجاری الماس کریمخان، طبقه دوم، واحد ۲۰۸",
                "landline_phone": "02186091015",
            },
        ]

        for dd in dealers_data:
            # Split full_name into first/last
            name_parts = dd["full_name"].split(" ", 1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ""

            existing = db.query(User).filter(User.mobile == dd["mobile"]).first()
            if not existing:
                dealer = User(
                    mobile=dd["mobile"],
                    first_name=first_name,
                    last_name=last_name,
                    national_id=dd["national_id"],
                    api_key=dd.get("api_key"),
                    tier_id=dd.get("tier_id"),
                    province_id=dd.get("province_id"),
                    city_id=dd.get("city_id"),
                    dealer_address=dd.get("address"),
                    landline_phone=dd.get("landline_phone"),
                    is_warehouse=dd.get("is_warehouse", False),
                    is_postal_hub=dd.get("is_postal_hub", False),
                    is_dealer=True,
                    is_customer=True,  # dealers can also shop
                )
                db.add(dealer)
                db.flush()
                tags = []
                if dd.get("is_warehouse"): tags.append("warehouse")
                if dd.get("is_postal_hub"): tags.append("postal_hub")
                tag_str = f" [{', '.join(tags)}]" if tags else ""
                print(f"  + {dd['full_name']}: {dd['mobile']}{tag_str}")
            else:
                # Ensure dealer flags are set
                if not existing.is_dealer:
                    existing.is_dealer = True
                if not existing.tier_id and dd.get("tier_id"):
                    existing.tier_id = dd["tier_id"]
                    print(f"  ~ updated tier: {dd['mobile']}")
                elif not existing.api_key and dd.get("api_key"):
                    existing.api_key = dd["api_key"]
                    print(f"  ~ updated api_key: {dd['mobile']}")
                else:
                    print(f"  = exists: {dd['mobile']}")

        db.flush()

        # ==========================================
        # 10.5. Bars (inventory) — now that dealers exist
        # ==========================================
        print("\n[10.5] Bars (Inventory)")

        existing_bar_count = db.query(Bar).count()
        if existing_bar_count > 0:
            print(f"  = {existing_bar_count} bars already exist, skipping")
        else:
            batch1 = _bar_batch1

            # Use all active dealers as locations for bars
            all_dealer_locs = db.query(User).filter(User.is_dealer == True, User.is_active == True).all()

            total_bars = 0
            used_serials = set()

            all_products = db.query(Product).filter(Product.is_active == True).all()
            for product in all_products:
                w = float(product.weight)
                if w <= 1:
                    count_per_dealer = 5
                elif w <= 5:
                    count_per_dealer = 3
                elif w <= 50:
                    count_per_dealer = 2
                else:
                    count_per_dealer = 1

                for dlr in all_dealer_locs:
                    for _ in range(count_per_dealer):
                        serial = generate_serial()
                        while serial in used_serials:
                            serial = generate_serial()
                        used_serials.add(serial)

                        db.add(Bar(
                            serial_code=serial,
                            status=BarStatus.ASSIGNED,
                            product_id=product.id,
                            batch_id=batch1.id if batch1 else None,
                            dealer_id=dlr.id,
                        ))
                        total_bars += 1

            db.flush()
            print(f"  + {total_bars} bars for {len(all_products)} products across {len(all_dealer_locs)} dealers")

            # --- Test bars for claim & transfer testing ---
            first_product = db.query(Product).first()
            first_dealer = db.query(User).filter(User.is_dealer == True, User.is_active == True).first()
            test_customer_1 = db.query(User).filter(User.mobile == "09351234567").first()

            if first_product and first_dealer:
                _now = datetime.now(timezone.utc)
                claim_bar_1 = Bar(
                    serial_code="TSCLM001", status=BarStatus.SOLD,
                    product_id=first_product.id, batch_id=batch1.id if batch1 else None,
                    dealer_id=first_dealer.id, customer_id=None, claim_code="ABC123",
                    delivered_at=_now,
                )
                db.add(claim_bar_1)
                claim_bar_2 = Bar(
                    serial_code="TSCLM002", status=BarStatus.SOLD,
                    product_id=first_product.id, batch_id=batch1.id if batch1 else None,
                    dealer_id=first_dealer.id, customer_id=None, claim_code="XYZ789",
                    delivered_at=_now,
                )
                db.add(claim_bar_2)
                if test_customer_1:
                    transfer_bar = Bar(
                        serial_code="TSTRF001", status=BarStatus.SOLD,
                        product_id=first_product.id, batch_id=batch1.id if batch1 else None,
                        dealer_id=first_dealer.id, customer_id=test_customer_1.id, claim_code=None,
                        delivered_at=_now,
                    )
                    db.add(transfer_bar)
                db.flush()
                print("  + 3 test bars (TSCLM001, TSCLM002, TSTRF001)")

            # --- Custodial test bars with orders (for /admin/orders page) ---
            test_customer_2 = db.query(User).filter(User.mobile == "09359876543").first()

            # Find a gold product and a silver product
            gold_product = (
                db.query(Product)
                .join(ProductCategoryLink, ProductCategoryLink.product_id == Product.id)
                .join(ProductCategory, ProductCategory.id == ProductCategoryLink.category_id)
                .filter(ProductCategory.slug.like("gold%"), Product.is_active == True)
                .first()
            )
            silver_product = (
                db.query(Product)
                .join(ProductCategoryLink, ProductCategoryLink.product_id == Product.id)
                .join(ProductCategory, ProductCategory.id == ProductCategoryLink.category_id)
                .filter(ProductCategory.slug.like("silver%"), Product.is_active == True)
                .first()
            )

            if gold_product and first_dealer and test_customer_1:
                _now = datetime.now(timezone.utc)
                gold_price = 52_000_000  # 5.2M toman/gram = 52M rial/gram
                tax_pct = 10

                # --- Custodial Gold Bar 1: regular (buyer = current owner) ---
                cust_bar_g1 = Bar(
                    serial_code="TSCST001", status=BarStatus.SOLD,
                    product_id=gold_product.id, batch_id=batch1.id if batch1 else None,
                    dealer_id=first_dealer.id, customer_id=test_customer_1.id,
                    delivered_at=None,  # custodial!
                )
                db.add(cust_bar_g1)
                db.flush()

                weight_g1 = float(gold_product.weight or 1)
                purity_g1 = float(gold_product.purity or 750)
                wage_g1 = float(gold_product.wage or 7)
                raw_gold_g1 = int(weight_g1 * (purity_g1 / 750) * gold_price)
                wage_amt_g1 = int(raw_gold_g1 * wage_g1 / 100)
                tax_amt_g1 = int(wage_amt_g1 * tax_pct / 100)
                line_total_g1 = raw_gold_g1 + wage_amt_g1 + tax_amt_g1

                order_g1 = Order(
                    customer_id=test_customer_1.id,
                    total_amount=line_total_g1,
                    status=OrderStatus.PAID,
                    delivery_method=DeliveryMethod.PICKUP,
                    delivery_status=DeliveryStatus.WAITING,
                    pickup_dealer_id=first_dealer.id,
                    payment_method="wallet",
                    paid_at=_now,
                )
                db.add(order_g1)
                db.flush()

                db.add(OrderItem(
                    order_id=order_g1.id, product_id=gold_product.id, bar_id=cust_bar_g1.id,
                    applied_metal_price=gold_price, applied_unit_price=line_total_g1,
                    applied_weight=weight_g1, applied_purity=purity_g1,
                    applied_wage_percent=wage_g1, applied_tax_percent=tax_pct,
                    final_gold_amount=raw_gold_g1, final_wage_amount=wage_amt_g1,
                    final_tax_amount=tax_amt_g1, line_total=line_total_g1,
                ))

                # --- Custodial Gold Bar 2: ownership transferred (buyer ≠ current owner) ---
                cust_bar_g2 = Bar(
                    serial_code="TSCST002", status=BarStatus.SOLD,
                    product_id=gold_product.id, batch_id=batch1.id if batch1 else None,
                    dealer_id=first_dealer.id,
                    customer_id=test_customer_2.id if test_customer_2 else test_customer_1.id,
                    delivered_at=None,  # custodial!
                )
                db.add(cust_bar_g2)
                db.flush()

                order_g2 = Order(
                    customer_id=test_customer_1.id,  # original buyer = customer 1
                    total_amount=line_total_g1,
                    status=OrderStatus.PAID,
                    delivery_method=DeliveryMethod.PICKUP,
                    delivery_status=DeliveryStatus.WAITING,
                    pickup_dealer_id=first_dealer.id,
                    payment_method="wallet",
                    paid_at=_now,
                )
                db.add(order_g2)
                db.flush()

                db.add(OrderItem(
                    order_id=order_g2.id, product_id=gold_product.id, bar_id=cust_bar_g2.id,
                    applied_metal_price=gold_price, applied_unit_price=line_total_g1,
                    applied_weight=weight_g1, applied_purity=purity_g1,
                    applied_wage_percent=wage_g1, applied_tax_percent=tax_pct,
                    final_gold_amount=raw_gold_g1, final_wage_amount=wage_amt_g1,
                    final_tax_amount=tax_amt_g1, line_total=line_total_g1,
                ))

                if test_customer_2:
                    db.add(OwnershipHistory(
                        bar_id=cust_bar_g2.id,
                        previous_owner_id=test_customer_1.id,
                        new_owner_id=test_customer_2.id,
                        description="انتقال مالکیت — تست",
                    ))

                db.flush()
                print(f"  + 2 custodial gold bars (TSCST001, TSCST002) with orders")

            # --- Custodial Silver Bar ---
            if silver_product and first_dealer and test_customer_1:
                _now = datetime.now(timezone.utc)
                silver_price = 1_500_000  # rial/gram
                tax_pct = 10

                cust_bar_s1 = Bar(
                    serial_code="TSCST003", status=BarStatus.SOLD,
                    product_id=silver_product.id, batch_id=batch1.id if batch1 else None,
                    dealer_id=first_dealer.id, customer_id=test_customer_1.id,
                    delivered_at=None,  # custodial!
                )
                db.add(cust_bar_s1)
                db.flush()

                weight_s1 = float(silver_product.weight or 1)
                purity_s1 = float(silver_product.purity or 999)
                wage_s1 = float(silver_product.wage or 5)
                raw_silver_s1 = int(weight_s1 * silver_price)
                wage_amt_s1 = int(raw_silver_s1 * wage_s1 / 100)
                tax_amt_s1 = int(wage_amt_s1 * tax_pct / 100)
                line_total_s1 = raw_silver_s1 + wage_amt_s1 + tax_amt_s1

                order_s1 = Order(
                    customer_id=test_customer_1.id,
                    total_amount=line_total_s1,
                    status=OrderStatus.PAID,
                    delivery_method=DeliveryMethod.PICKUP,
                    delivery_status=DeliveryStatus.WAITING,
                    pickup_dealer_id=first_dealer.id,
                    payment_method="wallet",
                    paid_at=_now,
                )
                db.add(order_s1)
                db.flush()

                db.add(OrderItem(
                    order_id=order_s1.id, product_id=silver_product.id, bar_id=cust_bar_s1.id,
                    applied_metal_price=silver_price, applied_unit_price=line_total_s1,
                    applied_weight=weight_s1, applied_purity=purity_s1,
                    applied_wage_percent=wage_s1, applied_tax_percent=tax_pct,
                    final_gold_amount=raw_silver_s1, final_wage_amount=wage_amt_s1,
                    final_tax_amount=tax_amt_s1, line_total=line_total_s1,
                ))
                db.flush()
                print(f"  + 1 custodial silver bar (TSCST003) with order")

        # Create dealer wallets (IRR + XAU_MG)
        print("\n  Dealer Wallets:")
        all_dealers = db.query(User).filter(User.is_dealer == True).all()
        for d in all_dealers:
            for asset in ["IRR", "XAU_MG"]:
                existing = db.query(Account).filter(
                    Account.user_id == d.id,
                    Account.asset_code == asset,
                ).first()
                if not existing:
                    db.add(Account(
                        user_id=d.id,
                        asset_code=asset,
                        balance=0, locked_balance=0,
                    ))
                    print(f"    + {d.full_name} ({asset})")
        db.flush()

        # ==========================================
        # 11. Sample Support Tickets
        # ==========================================
        print("\n[11] Sample Tickets")

        existing_tickets = db.query(Ticket).count()
        if existing_tickets == 0:
            test_customer_1 = db.query(User).filter(User.mobile == "09351234567").first()
            test_dealer_1 = db.query(User).filter(User.mobile == "09161234567", User.is_dealer == True).first()
            admin_user = db.query(User).filter(User.mobile == "09123456789", User.is_admin == True).first()

            if test_customer_1:
                t1 = Ticket(
                    subject="مشکل در پرداخت آنلاین",
                    body="سلام، من سعی کردم از درگاه بانکی پرداخت کنم ولی بعد از رفتن به صفحه بانک، خطای اتصال دریافت می‌کنم. لطفا بررسی کنید.",
                    sender_type=SenderType.CUSTOMER,
                    user_id=test_customer_1.id,
                    priority=TicketPriority.HIGH,
                    category=TicketCategory.FINANCIAL,
                    status=TicketStatus.ANSWERED,
                    assigned_to=admin_user.id if admin_user else None,
                )
                db.add(t1)
                db.flush()
                db.add(TicketMessage(
                    ticket_id=t1.id,
                    sender_type=SenderType.CUSTOMER,
                    sender_name=test_customer_1.full_name,
                    body=t1.body,
                    is_initial=True,
                ))
                db.add(TicketMessage(
                    ticket_id=t1.id,
                    sender_type=SenderType.STAFF,
                    sender_name="مدیر سیستم",
                    body="سلام، مشکل درگاه بانکی بررسی و رفع شد. لطفا مجددا تلاش کنید.",
                ))
                print(f"  + Ticket #{t1.id}: {t1.subject} [Financial]")

                t2 = Ticket(
                    subject="سوال درباره گارانتی شمش",
                    body="آیا شمش‌های خریداری شده گارانتی اصالت دارند؟ چطور می‌توانم اصالت شمش را بررسی کنم؟",
                    sender_type=SenderType.CUSTOMER,
                    user_id=test_customer_1.id,
                    priority=TicketPriority.LOW,
                    category=TicketCategory.SALES,
                    status=TicketStatus.OPEN,
                )
                db.add(t2)
                db.flush()
                db.add(TicketMessage(
                    ticket_id=t2.id,
                    sender_type=SenderType.CUSTOMER,
                    sender_name=test_customer_1.full_name,
                    body=t2.body,
                    is_initial=True,
                ))
                print(f"  + Ticket #{t2.id}: {t2.subject} [Sales]")

            if test_dealer_1:
                t3 = Ticket(
                    subject="درخواست افزایش موجودی شعبه",
                    body="موجودی شمش‌های ۱ گرمی در شعبه اصفهان تمام شده. لطفا موجودی جدید ارسال کنید.",
                    sender_type=SenderType.DEALER,
                    user_id=test_dealer_1.id,
                    priority=TicketPriority.MEDIUM,
                    category=TicketCategory.SALES,
                    status=TicketStatus.IN_PROGRESS,
                    assigned_to=admin_user.id if admin_user else None,
                )
                db.add(t3)
                db.flush()
                db.add(TicketMessage(
                    ticket_id=t3.id,
                    sender_type=SenderType.DEALER,
                    sender_name=test_dealer_1.full_name,
                    body=t3.body,
                    is_initial=True,
                ))
                db.add(TicketMessage(
                    ticket_id=t3.id,
                    sender_type=SenderType.STAFF,
                    sender_name="مدیر سیستم",
                    body="درخواست شما ثبت شد. ارسال از انبار مرکزی در ۲ روز کاری انجام خواهد شد.",
                ))
                db.add(TicketMessage(
                    ticket_id=t3.id,
                    sender_type=SenderType.DEALER,
                    sender_name=test_dealer_1.full_name,
                    body="متشکرم. لطفا بعد از ارسال کد رهگیری پست را هم اطلاع دهید.",
                ))
                print(f"  + Ticket #{t3.id}: {t3.subject} [Sales]")

            db.flush()
        else:
            print(f"  = {existing_tickets} tickets exist, skipping")

        # ==========================================
        # 12. Sample Reviews & Comments
        # ==========================================
        print("\n[12] Sample Reviews & Comments")

        existing_reviews = db.query(Review).count()
        if existing_reviews == 0:
            test_customer_1 = db.query(User).filter(User.mobile == "09351234567").first()
            test_customer_2 = db.query(User).filter(User.mobile == "09359876543").first()
            first_product = db.query(Product).filter(Product.is_active == True).first()
            second_product = db.query(Product).filter(
                Product.is_active == True,
                Product.id != (first_product.id if first_product else 0),
            ).first()

            if test_customer_1 and first_product:
                # Find a paid order item for this customer+product (if exists)
                test_order_item = db.query(OrderItem).join(Order).filter(
                    Order.customer_id == test_customer_1.id,
                    Order.status == OrderStatus.PAID,
                    OrderItem.product_id == first_product.id,
                ).first()

                r1 = Review(
                    product_id=first_product.id,
                    user_id=test_customer_1.id,
                    order_item_id=test_order_item.id if test_order_item else None,
                    rating=5,
                    body="کیفیت شمش عالی بود و بسته‌بندی هم بسیار مناسب. حتما دوباره خرید می‌کنم.",
                    admin_reply="ممنون از اعتماد شما. خوشحالیم که رضایت داشتید.",
                    admin_reply_at=datetime.now(timezone.utc),
                )
                db.add(r1)
                db.flush()
                print(f"  + Review #{r1.id}: {first_product.name} (5 stars)")

                if second_product:
                    r2 = Review(
                        product_id=second_product.id,
                        user_id=test_customer_1.id,
                        rating=4,
                        body="محصول خوبی بود. فقط زمان ارسال کمی طولانی شد.",
                    )
                    db.add(r2)
                    db.flush()
                    print(f"  + Review #{r2.id}: {second_product.name} (4 stars)")

            if test_customer_2 and first_product:
                r3 = Review(
                    product_id=first_product.id,
                    user_id=test_customer_2.id,
                    rating=3,
                    body="شمش اصل بود ولی قیمت نسبت به بازار کمی بالاتر بود.",
                )
                db.add(r3)
                db.flush()
                print(f"  + Review #{r3.id}: {first_product.name} (3 stars)")

            # Comments (Q&A on product page)
            if test_customer_1 and first_product:
                c1 = ProductComment(
                    product_id=first_product.id,
                    user_id=test_customer_1.id,
                    sender_name=test_customer_1.full_name or "مشتری",
                    sender_type="CUSTOMER",
                    body="آیا این شمش قابلیت ضرب سفارشی دارد؟",
                )
                db.add(c1)
                db.flush()

                # Admin reply to comment
                c1_reply = ProductComment(
                    product_id=first_product.id,
                    user_id=None,
                    sender_name="پشتیبانی",
                    sender_type="ADMIN",
                    parent_id=c1.id,
                    body="سلام، بله امکان ضرب سفارشی برای سفارشات بالای ۱۰ عدد وجود دارد. با پشتیبانی تماس بگیرید.",
                )
                db.add(c1_reply)
                db.flush()
                print(f"  + Comment #{c1.id}: Q&A with admin reply")

            if test_customer_2 and first_product:
                c2 = ProductComment(
                    product_id=first_product.id,
                    user_id=test_customer_2.id,
                    sender_name=test_customer_2.full_name or "مشتری",
                    sender_type="CUSTOMER",
                    body="زمان تحویل حضوری چقدر طول می‌کشد؟",
                )
                db.add(c2)
                db.flush()
                print(f"  + Comment #{c2.id}: pending question")

                # Like on comment
                if test_customer_1:
                    db.add(CommentLike(comment_id=c2.id, user_id=test_customer_1.id))

            db.flush()
        else:
            print(f"  = {existing_reviews} reviews exist, skipping")

        # ==========================================
        # Commit
        # ==========================================
        db.commit()

        print("\n" + "=" * 50)
        print("  Seed completed successfully!")
        print("=" * 50)

        print("\n--- Summary ---")
        print(f"  Total Users:    {db.query(User).count()}")
        print(f"    Admins:       {db.query(User).filter(User.is_admin == True).count()}")
        print(f"    Customers:    {db.query(User).filter(User.is_customer == True).count()}")
        print(f"    Dealers:      {db.query(User).filter(User.is_dealer == True).count()}")
        print(f"  Settings:       {db.query(SystemSetting).count()}")
        print(f"  Products:       {db.query(Product).count()}")
        print(f"  Categories:     {db.query(ProductCategory).count()}")
        print(f"  Provinces:      {db.query(GeoProvince).count()}")
        print(f"  Cities:         {db.query(GeoCity).count()}")
        print(f"  Card Designs:   {db.query(CardDesign).count()}")
        print(f"  Package Types:  {db.query(PackageType).count()}")
        print(f"  Batches:        {db.query(Batch).count()}")
        print(f"  Bars:           {db.query(Bar).count()}")
        print(f"  Orders:         {db.query(Order).count()}")
        print(f"  Coupons:        {db.query(Coupon).count()}")
        print(f"  Wallet Accts:   {db.query(Account).count()}")
        print(f"  Dealer Tiers:   {db.query(DealerTier).count()}")
        print(f"  Tier Wages:     {db.query(ProductTierWage).count()}")
        print(f"  Tickets:        {db.query(Ticket).count()}")
        print(f"  Reviews:        {db.query(Review).count()}")
        print(f"  Comments:       {db.query(ProductComment).count()}")
        print(f"  Withdrawals:    {db.query(WithdrawalRequest).count()}")

        print("\n--- Credentials ---")
        print(f"  Admin:    09123456789")
        print(f"  Operator: 09121111111")
        print(f"  Customer: 09351234567 (wallet: 10M toman)")
        print(f"  Customer: 09359876543")
        print(f"  Customer: 09131112233")
        print(f"  Dealer:   09161234567 (esfahan, tier: distributor)")
        print(f"  Dealer:   09171234567 (shiraz, tier: wholesaler)")
        print(f"  Dealer:   09181234567 (mashhad, tier: store)")

        print(f"\n--- Dealer API Keys (for POS) ---")
        for dd in dealers_data:
            print(f"  {dd['mobile']}: {dd['api_key']}")

        print(f"\n--- Custodial Test Bars ---")
        print(f"  TSCST001 : Gold, custodial (buyer=owner=09351234567)")
        print(f"  TSCST002 : Gold, custodial (buyer=09351234567, owner=09359876543 — transferred)")
        print(f"  TSCST003 : Silver, custodial (buyer=owner=09351234567)")

        print(f"\n--- Settings ---")
        print(f"  Gold price: 5,200,000 toman/gram")
        print(f"  Tax: 10%")
        print(f"  Shipping: 50,000 toman")

        print(f"\n--- Coupons ---")
        print(f"  WELCOME10 : 10% off (first purchase)")
        print(f"  CASHBACK5 : 5% cashback")
        print(f"  FIXED500  : 500K toman off (min 5M order)")
        print(f"  VIP2026   : 15% off (mobile: 09351234567, 09359876543)")
        print(f"  GOLD10    : 10% off (category: gold-talamala + gold-investment)")

    except Exception as e:
        db.rollback()
        print(f"\nSeed failed: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


def reset_and_seed():
    """Drop all tables and recreate + seed."""
    from sqlalchemy import text
    with engine.connect() as conn:
        # Drop legacy/renamed tables that may have FK constraints
        conn.execute(text("DROP TABLE IF EXISTS location_transfers CASCADE"))
        conn.execute(text("DROP TABLE IF EXISTS package_images CASCADE"))
        conn.execute(text("DROP TABLE IF EXISTS design_images CASCADE"))
        # Drop old separate user tables (merged into 'users')
        conn.execute(text("DROP TABLE IF EXISTS customers CASCADE"))
        conn.execute(text("DROP TABLE IF EXISTS dealers CASCADE"))
        conn.execute(text("DROP TABLE IF EXISTS system_users CASCADE"))
        conn.commit()
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    print("All tables recreated")
    seed()


if __name__ == "__main__":
    if "--reset" in sys.argv:
        confirm = input("This will DELETE ALL DATA. Type 'yes' to confirm: ")
        if confirm.strip().lower() == "yes":
            reset_and_seed()
        else:
            print("Aborted.")
    else:
        seed()
