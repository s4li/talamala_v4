#!/usr/bin/env python3
"""
Import bar serials from an Excel workbook.

Each sheet name is a product_id; the `serial` column holds the serial codes.
Imported bars are created as Sold with no owner and no dealer location — they
are historical bars already out in the world, recorded so they resolve on the
public authenticity check. They are never sellable.

Runs as a DRY RUN unless --commit is passed. The dry run touches nothing and
prints exactly what a real run would do, so always run it first.

    python scripts/import_serials.py                      # dry run
    python scripts/import_serials.py --commit             # actually write
    python scripts/import_serials.py -f other.xlsx --commit

Everything is written in a single transaction: if any row fails, nothing is
saved.
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import main  # noqa: F401  — registers every model with SQLAlchemy
from config.database import SessionLocal
from modules.catalog.models import Product
from modules.inventory.models import Bar, BarStatus
from modules.inventory.service import SAFE_CHARS

SERIAL_MAX_LEN = 8


def parse_args():
    p = argparse.ArgumentParser(description="Import bar serials from Excel into the bars table.")
    p.add_argument("-f", "--file", default="serials.xlsx", help="Excel file (default: serials.xlsx)")
    p.add_argument("-c", "--column", default="serial", help="Serial column header (default: serial)")
    p.add_argument("--status", default=BarStatus.SOLD.value,
                   choices=[s.value for s in BarStatus],
                   help="Status for imported bars (default: Sold)")
    p.add_argument("--commit", action="store_true",
                   help="Actually write. Without it the script only reports.")
    return p.parse_args()


def read_workbook(path, column):
    """-> ({sheet_name: [serial, ...]}, [warning, ...])   Raises on a bad file."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("!! openpyxl is not installed.  pip install openpyxl")

    if not os.path.exists(path):
        sys.exit(f"!! file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets, warnings = {}, []

    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            warnings.append(f"sheet {name!r}: empty, skipped")
            continue

        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        if column.lower() not in header:
            warnings.append(f"sheet {name!r}: no {column!r} column (found: {header}), skipped")
            continue
        idx = header.index(column.lower())

        serials = []
        for row_no, row in enumerate(rows[1:], start=2):
            if idx >= len(row) or row[idx] is None:
                continue
            value = str(row[idx]).strip().upper()
            if not value:
                continue
            serials.append((row_no, value))
        sheets[name] = serials

    return sheets, warnings


def main_import():
    args = parse_args()
    sheets, warnings = read_workbook(args.file, args.column)

    print(f"file:   {args.file}")
    print(f"status: {args.status}")
    print(f"mode:   {'COMMIT — writing to the database' if args.commit else 'DRY RUN — nothing will be written'}")
    print()
    for w in warnings:
        print(f"  warn: {w}")
    if warnings:
        print()

    if not sheets:
        sys.exit("!! nothing to import")

    db = SessionLocal()
    errors, to_create, skipped = [], [], []
    seen_in_file = {}

    for sheet_name, rows in sheets.items():
        # Sheet name is the product id
        if not str(sheet_name).strip().isdigit():
            errors.append(f"sheet {sheet_name!r}: name is not a product_id")
            continue
        product_id = int(str(sheet_name).strip())

        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            errors.append(f"sheet {sheet_name!r}: product_id {product_id} does not exist")
            continue

        print(f"sheet {sheet_name!r} -> product {product_id} ({product.name}): {len(rows)} rows")

        for row_no, serial in rows:
            where = f"sheet {sheet_name!r} row {row_no} ({serial})"

            if len(serial) > SERIAL_MAX_LEN:
                errors.append(f"{where}: longer than {SERIAL_MAX_LEN} characters")
                continue
            bad = {c for c in serial if c not in SAFE_CHARS}
            if bad:
                errors.append(f"{where}: invalid characters {''.join(sorted(bad))}")
                continue
            if serial in seen_in_file:
                errors.append(f"{where}: duplicated in file, first seen at {seen_in_file[serial]}")
                continue
            seen_in_file[serial] = where

            if db.query(Bar).filter(Bar.serial_code == serial).first():
                skipped.append(f"{where}: already in the database")
                continue

            to_create.append((serial, product_id))

    print()
    for s in skipped:
        print(f"  skip: {s}")
    for e in errors:
        print(f"  ERROR: {e}")

    print()
    print(f"to create: {len(to_create)}    already present: {len(skipped)}    errors: {len(errors)}")

    if errors:
        db.close()
        sys.exit("\n!! errors found — nothing was written. Fix the file and run again.")

    if not to_create:
        db.close()
        print("\nnothing new to import.")
        return

    if not args.commit:
        db.close()
        print("\nDRY RUN — nothing written. Re-run with --commit to apply.")
        return

    for serial, product_id in to_create:
        db.add(Bar(
            serial_code=serial,
            status=args.status,
            product_id=product_id,
            customer_id=None,    # historical sale, owner unknown
            dealer_id=None,      # no longer in our warehouse
            is_sellable=False,   # sold bars are never offered for sale
        ))

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        db.close()
        sys.exit(f"\n!! commit failed, nothing was written: {exc}")

    total = db.query(Bar).count()
    db.close()
    print(f"\nDONE — {len(to_create)} bars created. bars table now holds {total} rows.")


if __name__ == "__main__":
    main_import()
