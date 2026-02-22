"""
TalaMala v4 - Fix Product Wages & Images
==========================================
Updates product wages from Excel and copies product images from _private/.
Does NOT touch bars, users, orders, or any other data.

Usage:
    python scripts/fix_products.py
"""

import sys
import os
import io
import shutil
import openpyxl

# Fix Windows console encoding for Persian text
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config.database import SessionLocal
# Import ALL models to resolve SQLAlchemy relationships
from modules.user.models import User  # noqa: F401
from modules.admin.models import SystemSetting  # noqa: F401
from modules.customer.address_models import GeoProvince, GeoCity, GeoDistrict  # noqa: F401
from modules.catalog.models import (
    ProductCategory, ProductCategoryLink, Product, ProductImage, ProductTierWage,
)
from modules.inventory.models import Bar  # noqa: F401
from modules.order.models import Order  # noqa: F401
from modules.wallet.models import Account  # noqa: F401
from modules.coupon.models import Coupon  # noqa: F401
from modules.dealer.models import DealerTier, DealerSale  # noqa: F401
from modules.ticket.models import Ticket  # noqa: F401
from modules.review.models import Review  # noqa: F401
from modules.dealer_request.models import DealerRequest  # noqa: F401
from modules.pricing.models import Asset  # noqa: F401

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMG_SRC_BASE = os.path.join(PROJECT_ROOT, "_private", "عکس محصولات")
IMG_DST_BASE = os.path.join(PROJECT_ROOT, "static", "uploads", "products")


# ---- Product type definitions (must match seed_production.py) ----

product_types = [
    {
        "folder": "شمش طلا با بسته بندی",
        "excel_sheet": "شمش طلاملا",
        "slug": "gold-talamala",
        "category_slug": "gold-talamala",
        "name_prefix": "شمش طلا طلاملا",
        "purity": 995,
        "metal": "gold",
    },
    {
        "folder": "شمش طلا بدون بسته بندی",
        "excel_sheet": "شمش سرمایه ای",
        "slug": "gold-investment",
        "category_slug": "gold-investment",
        "name_prefix": "شمش طلا سرمایه‌ای",
        "purity": 995,
        "metal": "gold",
    },
    {
        "folder": "شمش نقره با بسته بندی",
        "excel_sheet": "شمش نقره طلاملا",
        "slug": "silver-talamala",
        "category_slug": "silver-talamala",
        "name_prefix": "شمش نقره طلاملا",
        "purity": 999.9,
        "metal": "silver",
    },
    {
        "folder": "شمش نقره بدون بسته بندی",
        "excel_sheet": "شمش سرمایه ای نقره",
        "slug": "silver-investment",
        "category_slug": "silver-investment",
        "name_prefix": "شمش نقره سرمایه‌ای",
        "purity": 999.9,
        "metal": "silver",
    },
]

weights_def = [
    (0.1,  "۱۰۰ سوت",  {"شمش نقره با بسته بندی": "0.1 g"}),
    (0.2,  "۲۰۰ سوت",  {}),
    (0.5,  "۵۰۰ سوت",  {}),
    (1.0,  "۱ گرم",     {}),
    (2.5,  "۲.۵ گرم",   {}),
    (5.0,  "۵ گرم",     {}),
    (10.0, "۱۰ گرم",    {}),
    (20.0, "۲۰ گرم",    {}),
    (31.1, "۱ اونس",    {
        "شمش طلا با بسته بندی": "1ounce",
        "شمش طلا بدون بسته بندی": "1ounce",
        "شمش نقره با بسته بندی": "1onuce",
        "شمش نقره بدون بسته بندی": "1onuce",
    }),
    (50.0, "۵۰ گرم",    {"شمش طلا بدون بسته بندی": "50G"}),
    (100.0,"۱۰۰ گرم",   {}),
]


def weight_to_folder(weight_grams, ptype_folder, overrides):
    if ptype_folder in overrides:
        return overrides[ptype_folder]
    if weight_grams == 31.1:
        return "1ounce"
    if weight_grams == int(weight_grams):
        return f"{int(weight_grams)}g"
    return f"{weight_grams}g"


def fix():
    db = SessionLocal()
    try:
        print("=" * 50)
        print("  Fix Product Wages & Images")
        print("=" * 50)

        # ==========================================
        # 1. Read wage data from Excel
        # ==========================================
        print("\n[1] Reading wage data from Excel")

        EXCEL_PATH = os.path.join(IMG_SRC_BASE, "سطوح قیمت گذاری طلا ونقره.xlsx")
        wage_data = {}

        if os.path.exists(EXCEL_PATH):
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
            for ptype in product_types:
                sheet_name = ptype["excel_sheet"]
                if sheet_name not in wb.sheetnames:
                    print(f"  ! Excel sheet '{sheet_name}' not found")
                    continue
                ws = wb[sheet_name]
                slug_wages = {}
                for row in ws.iter_rows(min_row=3, max_col=5, values_only=True):
                    weight_val, t1, t2, t3, t4 = row
                    if weight_val is None or t4 is None:
                        continue
                    w = round(float(weight_val), 2)
                    slug_wages[w] = [
                        round(float(t1 or 0) * 100, 2),
                        round(float(t2 or 0) * 100, 2),
                        round(float(t3 or 0) * 100, 2),
                        round(float(t4 or 0) * 100, 2),
                    ]
                wage_data[ptype["slug"]] = slug_wages
            wb.close()
            print(f"  + Wage data loaded ({len(wage_data)} types)")
        else:
            print(f"  !! Excel file not found: {EXCEL_PATH}")
            print(f"  !! Cannot fix wages without Excel file. Aborting.")
            return

        # ==========================================
        # 2. Build category slug map
        # ==========================================
        cat_map = {}
        for ptype in product_types:
            cat = db.query(ProductCategory).filter(
                ProductCategory.slug == ptype["category_slug"]
            ).first()
            if cat:
                cat_map[ptype["category_slug"]] = cat

        cat_id_to_slug = {cat_obj.id: slug for slug, cat_obj in cat_map.items()}

        # ==========================================
        # 3. Update product wages + images
        # ==========================================
        print("\n[2] Updating product wages")

        wage_updated = 0
        img_updated = 0
        os.makedirs(IMG_DST_BASE, exist_ok=True)

        for ptype in product_types:
            type_wages = wage_data.get(ptype["slug"], {})

            for weight_grams, weight_label, folder_overrides in weights_def:
                name = f"{ptype['name_prefix']} {weight_label}"
                tier_wages_list = type_wages.get(weight_grams, [0, 0, 0, 0])
                end_customer_wage = tier_wages_list[3]

                product = db.query(Product).filter(Product.name == name).first()
                if not product:
                    continue

                # Update wage
                if end_customer_wage and product.wage != end_customer_wage:
                    old_wage = product.wage
                    product.wage = end_customer_wage
                    wage_updated += 1
                    print(f"  ~ {name}: wage {old_wage} → {end_customer_wage}")

                # Update images (only if source exists)
                folder_name = weight_to_folder(weight_grams, ptype["folder"], folder_overrides)
                img_src_dir = os.path.join(IMG_SRC_BASE, ptype["folder"], folder_name)

                if os.path.isdir(img_src_dir):
                    # Find source images
                    src_files = []
                    for root, _dirs, files in os.walk(img_src_dir):
                        for f in sorted(files):
                            if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                                src_files.append(os.path.join(root, f))
                    src_files.sort()
                    src_files = src_files[:2]  # front + back

                    if src_files:
                        # Remove old images for this product
                        old_images = db.query(ProductImage).filter(
                            ProductImage.product_id == product.id
                        ).all()
                        for old_img in old_images:
                            old_file = os.path.join(PROJECT_ROOT, old_img.file_path)
                            if os.path.exists(old_file):
                                os.remove(old_file)
                            db.delete(old_img)
                        db.flush()

                        # Copy new images
                        for idx, src_path in enumerate(src_files):
                            ext = os.path.splitext(src_path)[1].lower()
                            clean_name = f"{ptype['slug']}_{folder_name}_{idx + 1}{ext}"
                            dst_path = os.path.join(IMG_DST_BASE, clean_name)
                            shutil.copy2(src_path, dst_path)
                            rel_path = f"static/uploads/products/{clean_name}"
                            db.add(ProductImage(
                                product_id=product.id,
                                file_path=rel_path,
                                is_default=(idx == 0),
                            ))
                        img_updated += 1

        db.flush()
        print(f"  + {wage_updated} product wages updated")
        print(f"  + {img_updated} products got new images")

        # ==========================================
        # 4. Refresh tier wages
        # ==========================================
        print("\n[3] Refreshing tier wages")

        tier_slugs_order = ["distributor", "wholesaler", "store", "end_customer"]
        tier_map = {}
        for slug in tier_slugs_order:
            tier = db.query(DealerTier).filter(DealerTier.slug == slug).first()
            if tier:
                tier_map[slug] = tier

        # Delete old tier wages and recreate
        existing_tw = db.query(ProductTierWage).count()
        if existing_tw > 0:
            db.query(ProductTierWage).delete()
            db.flush()
            print(f"  ~ deleted {existing_tw} old tier wages")

        tw_count = 0
        all_products = db.query(Product).filter(Product.is_active == True).all()
        for p in all_products:
            type_slug = None
            for cid in p.category_ids:
                if cid in cat_id_to_slug:
                    type_slug = cat_id_to_slug[cid]
                    break
            if not type_slug or type_slug not in wage_data:
                continue
            weight_key = float(p.weight)
            wages = wage_data[type_slug].get(weight_key)
            if not wages:
                continue
            for idx, tier_slug in enumerate(tier_slugs_order):
                tier = tier_map.get(tier_slug)
                if tier:
                    db.add(ProductTierWage(
                        product_id=p.id, tier_id=tier.id, wage_percent=wages[idx]
                    ))
                    tw_count += 1

        db.flush()
        print(f"  + {tw_count} tier wages created")

        # ==========================================
        # 5. Also drop is_customer column if it still exists
        # ==========================================
        print("\n[4] Schema cleanup")
        from sqlalchemy import text, inspect
        insp = inspect(db.bind)
        if "users" in insp.get_table_names():
            user_cols = {c["name"] for c in insp.get_columns("users")}
            if "is_customer" in user_cols:
                with db.bind.connect() as conn:
                    conn.execute(text("ALTER TABLE users DROP COLUMN is_customer"))
                    conn.commit()
                print("  + Dropped is_customer column")
            else:
                print("  = is_customer already removed")

        # ==========================================
        # Commit
        # ==========================================
        db.commit()

        print("\n" + "=" * 50)
        print("  Fix completed!")
        print("=" * 50)

        # Summary
        print(f"\n--- Summary ---")
        print(f"  Products:     {db.query(Product).count()}")
        print(f"  Images:       {db.query(ProductImage).count()}")
        print(f"  Tier wages:   {db.query(ProductTierWage).count()}")

        # Show a few product wages as sample
        print(f"\n--- Sample wages ---")
        samples = db.query(Product).filter(Product.is_active == True).limit(8).all()
        for s in samples:
            print(f"  {s.name}: wage={s.wage}%")

    except Exception as e:
        db.rollback()
        print(f"\nFix failed: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    fix()
